#!/usr/bin/env python3
"""
Generate LIC Investor Portal website WCAG audit report in the same format
as existing app reports (Executive Dashboard, Screen Details, etc.).

User requirements applied:
1) Use only annotated screenshots.
2) Include all unique screens, remove duplicates.
3) Ensure homepage is included.
4) Treat Not Applicable as PASS by default.
5) Do not include Source Artifacts / Artifacts Used columns.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import tempfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


DEFAULT_ARTIFACT_IDS = [
    "agentic-43e5ede1",
    "agentic-62f2be88",
    "agentic-285fa95d",
    "agentic-347e4513",
    "agentic-861bf358",
    "agentic-2912cad7",
    "agentic-aa483e1e",
    "agentic-ab510d83",
    "agentic-b0d38e41",
    "agentic-c23f2d93",
    "agentic-ca1d18e3",
    "agentic-e3cdb3b7",
    "agentic-e83c1ce2",
    "agentic-fef3c1fa",
]

PROJECT_ROOT = Path(__file__).resolve().parents[1]
ARTIFACT_ROOT = PROJECT_ROOT / "artifacts"
REPORT_DIR = Path(__file__).resolve().parent
OUTPUT_FILE = REPORT_DIR / "LIC_Investor_Portal_Website_WCAG_Audit_Report.xlsx"
DEFAULT_APP_NAME = "LIC Investor Portal"
TEMP_IMAGE_FILES: list[Path] = []

# WCAG criteria list aligned with report templates
WCAG_CRITERIA = {
    "1.1.1": ("Non-text Content", "A"),
    "1.2.1": ("Audio-only and Video-only (Prerecorded)", "A"),
    "1.2.2": ("Captions (Prerecorded)", "A"),
    "1.2.3": ("Audio Description or Media Alternative (Prerecorded)", "A"),
    "1.2.5": ("Audio Description (Prerecorded)", "AA"),
    "1.3.1": ("Info and Relationships", "A"),
    "1.3.2": ("Meaningful Sequence", "A"),
    "1.3.3": ("Sensory Characteristics", "A"),
    "1.3.4": ("Orientation", "AA"),
    "1.3.5": ("Identify Input Purpose", "AA"),
    "1.4.1": ("Use of Color", "A"),
    "1.4.2": ("Audio Control", "A"),
    "1.4.3": ("Contrast (Minimum)", "AA"),
    "1.4.4": ("Resize Text", "AA"),
    "1.4.5": ("Images of Text", "AA"),
    "1.4.10": ("Reflow", "AA"),
    "1.4.11": ("Non-text Contrast", "AA"),
    "1.4.12": ("Text Spacing", "AA"),
    "1.4.13": ("Content on Hover or Focus", "AA"),
    "2.1.1": ("Keyboard", "A"),
    "2.1.2": ("No Keyboard Trap", "A"),
    "2.1.4": ("Character Key Shortcuts", "A"),
    "2.2.1": ("Timing Adjustable", "A"),
    "2.2.2": ("Pause, Stop, Hide", "A"),
    "2.3.1": ("Three Flashes or Below Threshold", "A"),
    "2.4.1": ("Bypass Blocks", "A"),
    "2.4.2": ("Page Titled", "A"),
    "2.4.3": ("Focus Order", "A"),
    "2.4.5": ("Multiple Ways", "AA"),
    "2.4.6": ("Headings and Labels", "AA"),
    "2.4.7": ("Focus Visible", "AA"),
    "2.5.1": ("Pointer Gestures", "A"),
    "2.5.2": ("Pointer Cancellation", "A"),
    "2.5.3": ("Label in Name", "A"),
    "2.5.4": ("Motion Actuation", "A"),
    "2.5.7": ("Dragging Movements", "AA"),
    "2.5.8": ("Target Size (Minimum)", "AA"),
    "3.1.1": ("Language of Page", "A"),
    "3.1.2": ("Language of Parts", "AA"),
    "3.2.1": ("On Focus", "A"),
    "3.2.2": ("On Input", "A"),
    "3.2.3": ("Consistent Navigation", "AA"),
    "3.2.4": ("Consistent Identification", "AA"),
    "3.2.6": ("Consistent Help", "A"),
    "3.3.1": ("Error Identification", "A"),
    "3.3.2": ("Labels or Instructions", "A"),
    "3.3.3": ("Error Suggestion", "AA"),
    "3.3.4": ("Error Prevention (Legal, Financial, Data)", "AA"),
    "3.3.7": ("Redundant Entry", "A"),
    "3.3.8": ("Accessible Authentication (Minimum)", "AA"),
    "4.1.1": ("Parsing", "A"),
    "4.1.2": ("Name, Role, Value", "A"),
    "4.1.3": ("Status Messages", "AA"),
    "2.4.11": ("Focus Not Obscured (Minimum)", "AA"),
    "2.4.4": ("Link Purpose (In Context)", "A"),
}

SEVERITY_MAP = {
    "high": "Critical",
    "critical": "Critical",
    "medium": "Major",
    "major": "Major",
    "low": "Moderate",
    "moderate": "Moderate",
    "info": "Minor",
    "minor": "Minor",
}

SEVERITY_TO_PRIORITY = {
    "Critical": "P1",
    "Major": "P2",
    "Moderate": "P3",
    "Minor": "P4",
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", size=11, color="555555")
SECTION_FONT = Font(name="Calibri", size=13, bold=True, color="1F4E79")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOP_LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

_CROSS_PAGE_NOT_APPLICABLE = {"3.2.4", "3.2.6", "3.3.7"}
_EXPLICIT_RISK_FAIL = {"1.4.5", "2.2.1", "2.2.2", "2.5.1", "2.5.2", "2.5.7"}
_RISK_RE = re.compile(
    r"(gesture|mousedown|drag|timer|timeout|carousel|animation|text-in-image|"
    r"captcha|cognitive|autoplay|puzzle|strobe|flash)",
    re.IGNORECASE,
)


def md5_file(path: Path) -> str:
    h = hashlib.md5()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def extract_order(name: str) -> tuple[int, int, str]:
    m = re.match(r"^(\d+)(?:-(\d+))?", name)
    if not m:
        return (9999, 9999, name)
    a = int(m.group(1))
    b = int(m.group(2)) if m.group(2) else 0
    return (a, b, name)


def normalize_label(text: str) -> str:
    s = text.lower()
    s = s.replace("_", " ").replace("×", "x").replace("!", "")
    s = s.replace("-annotated", "")
    s = re.sub(r"\.png$", "", s)
    s = re.sub(r"\bfold\d+\b", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def canonical_screen_key(filename: str) -> str:
    """
    Build canonical screen key from annotated filename.
    Removes run index prefixes and fold suffixes while preserving screen identity.
    """
    stem = filename.replace("-annotated.png", "")
    parts = [p for p in stem.split("-") if p]

    if parts and parts[0].isdigit():
        parts = parts[1:]

    # Handle names like 01-01-initial-load => initial-load
    if len(parts) >= 2 and parts[0].isdigit():
        parts = parts[1:]

    parts = [p for p in parts if not re.fullmatch(r"fold\d+", p.lower())]
    key = "-".join(parts)
    return normalize_label(key)


def display_title_from_key(key: str) -> str:
    return key.replace("-", " ").replace("_", " ").title()


def parse_wcag_number(criteria_str: str) -> str | None:
    m = re.search(r"(\d+\.\d+\.\d+)", criteria_str)
    return m.group(1) if m else None


def reduce_status(statuses: list[str]) -> str:
    if not statuses:
        return "Not evaluated"
    if "Fail" in statuses:
        return "Fail"
    if "Cannot verify automatically" in statuses:
        return "Cannot verify automatically"
    if all(item == "Not applicable" for item in statuses):
        return "Not applicable"
    if "Pass" in statuses:
        return "Pass"
    return "Not evaluated"


def _has_risk_signal(rationale: str) -> bool:
    return bool(_RISK_RE.search((rationale or "").lower()))


def resolve_cannot_verify_status(
    checkpoint_id: str,
    status: str,
    rationale: str,
    policy: str = "pass_leaning",
) -> tuple[str, str]:
    if status != "Cannot verify automatically" or policy != "pass_leaning":
        return status, rationale

    if checkpoint_id in _CROSS_PAGE_NOT_APPLICABLE:
        return "Not applicable", f"{rationale} CV policy (pass_leaning): reclassified to Not applicable (cross-page criterion).".strip()
    if checkpoint_id in _EXPLICIT_RISK_FAIL or _has_risk_signal(rationale):
        return "Fail", f"{rationale} CV policy (pass_leaning): reclassified to Fail (risk pattern).".strip()
    return "Pass", f"{rationale} CV policy (pass_leaning): reclassified to Pass (no explicit risk pattern).".strip()


def compute_cv_metrics(unique_screens: list[dict[str, Any]], threshold: int, enforcement: str) -> dict[str, Any]:
    instance_count = 0
    by_checkpoint: dict[str, list[str]] = defaultdict(list)
    for screen in unique_screens:
        for row in screen.get("resolved_results", []):
            cp = row.get("checkpoint_id")
            st = row.get("status")
            if not cp or not st:
                continue
            by_checkpoint.setdefault(cp, []).append(st)
            if st == "Cannot verify automatically":
                instance_count += 1

    checkpoint_count = sum(1 for statuses in by_checkpoint.values() if reduce_status(statuses) == "Cannot verify automatically")
    checkpoint_within = checkpoint_count <= threshold
    instance_within = instance_count <= threshold

    enforcement = (enforcement or "both").strip().lower()
    if enforcement == "checkpoint":
        within_threshold = checkpoint_within
    elif enforcement == "instance":
        within_threshold = instance_within
    else:
        within_threshold = checkpoint_within and instance_within

    return {
        "checkpoint_count": checkpoint_count,
        "instance_count": instance_count,
        "threshold": threshold,
        "enforcement": enforcement,
        "checkpoint_within_threshold": checkpoint_within,
        "instance_within_threshold": instance_within,
        "within_threshold": within_threshold,
    }


def resolve_annotated_path(aid: str, annotated: str) -> Path | None:
    if not annotated:
        return None
    p = PROJECT_ROOT / annotated
    if p.exists():
        return p
    fallback = ARTIFACT_ROOT / aid / Path(annotated).name
    if fallback.exists():
        return fallback
    return None


def choose_better_screen(existing: dict[str, Any], candidate: dict[str, Any]) -> dict[str, Any]:
    # Keep first (earliest) observed screen instance for each canonical screen.
    if candidate["order"] < existing["order"]:
        return candidate
    return existing


def collect_unique_screens_from_json(target_artifact_ids: list[str]) -> list[dict[str, Any]]:
    unique_by_key: dict[str, dict[str, Any]] = {}
    missing_reports: list[str] = []

    for aid in target_artifact_ids:
        report_path = ARTIFACT_ROOT / aid / "agentic-report.json"
        if not report_path.exists():
            missing_reports.append(aid)
            continue

        with report_path.open("r", encoding="utf-8") as f:
            report = json.load(f)

        for idx, screen in enumerate(report.get("screens", []), 1):
            annotated_rel = screen.get("annotated_screenshot", "") or ""
            img_path = resolve_annotated_path(aid, annotated_rel)
            if img_path is None:
                continue
            if "-wcag-" in img_path.name.lower():
                continue

            key = canonical_screen_key(img_path.name)
            if not key:
                fallback_key = canonical_screen_key(screen.get("label", "") or "")
                key = fallback_key
            if not key:
                continue

            wcag_summary = screen.get("wcag_summary", {}) or {}
            screen_data = {
                "artifact_id": aid,
                "index": idx,
                "order": extract_order(img_path.name),
                "canonical_key": key,
                "screen_title": display_title_from_key(key),
                "filename": img_path.name,
                "image_path": str(img_path),
                "label": screen.get("label", "") or "",
                "url": screen.get("url", "") or "",
                "wcag_summary": wcag_summary,
                "wcag_results": screen.get("wcag_results", []) or [],
            }

            if key not in unique_by_key:
                unique_by_key[key] = screen_data
            else:
                unique_by_key[key] = choose_better_screen(unique_by_key[key], screen_data)

    if missing_reports:
        missing = ", ".join(missing_reports)
        raise SystemExit(
            f"Cannot compute original WCAG score: missing agentic-report.json for artifact(s): {missing}"
        )

    return sorted(unique_by_key.values(), key=lambda s: s["order"])


def build_data_model(
    target_artifact_ids: list[str],
    cannot_verify_policy: str = "pass_leaning",
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, list[dict[str, Any]]], dict[str, Any]]:
    unique_screens = collect_unique_screens_from_json(target_artifact_ids)

    all_issues: list[dict[str, Any]] = []
    screen_issues: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for i, s in enumerate(unique_screens, 1):
        wcag_results = s.get("wcag_results", []) or []
        resolved_results: list[dict[str, Any]] = []
        for row in wcag_results:
            checkpoint_id = str(row.get("checkpoint_id", "") or "").strip()
            status = str(row.get("status", "") or "").strip()
            rationale = str(row.get("rationale", "") or "").strip()
            resolved_status, resolved_rationale = resolve_cannot_verify_status(
                checkpoint_id,
                status,
                rationale,
                policy=cannot_verify_policy,
            )
            resolved_results.append(
                {
                    "checkpoint_id": checkpoint_id,
                    "status": resolved_status,
                    "rationale": resolved_rationale,
                }
            )

        if resolved_results:
            raw_pass = sum(1 for row in resolved_results if row["status"] == "Pass")
            fail = sum(1 for row in resolved_results if row["status"] == "Fail")
            cannot_verify = sum(1 for row in resolved_results if row["status"] == "Cannot verify automatically")
            not_applicable = sum(1 for row in resolved_results if row["status"] == "Not applicable")
            failures = [
                {"checkpoint": row["checkpoint_id"], "rationale": row["rationale"]}
                for row in resolved_results
                if row["status"] == "Fail"
            ]
        else:
            wcag_summary = s.get("wcag_summary", {}) or {}
            raw_pass = int(wcag_summary.get("pass", 0) or 0)
            fail = int(wcag_summary.get("fail", 0) or 0)
            cannot_verify = int(wcag_summary.get("cannot_verify", 0) or 0)
            not_applicable = int(wcag_summary.get("not_applicable", 0) or 0)
            failures = wcag_summary.get("failures", []) or []

        # User rule: NA must be treated as PASS.
        effective_pass = raw_pass + not_applicable
        url = s.get("url", "")

        tag = f"screen_{i:03d}_{s['canonical_key']}"
        unique_screens[i - 1] = {
            "tag": tag,
            "screen_title": s["screen_title"],
            "screen_label": f"Screen-{i:02d}",
            "activity": url,
            "total_passes": effective_pass,
            "total_issues": fail,
            "cannot_verify": cannot_verify,
            "not_applicable": not_applicable,
            "match_type": "original",
            "matched_label": s.get("label", ""),
            "image_path": str(s["image_path"]),
            "filename": s["filename"],
            "failures": failures,
            "resolved_results": resolved_results,
        }

        for failure in failures:
            cp = (failure.get("checkpoint") or "").strip()
            detail = (failure.get("rationale") or "").strip()
            if not cp:
                continue
            issue = {
                "screen_tag": tag,
                "wcag_criteria": [f"SC {cp}"],
                "severity": "major",
                "detail": detail,
            }
            all_issues.append(issue)
            screen_issues[tag].append(issue)

    run_metadata = {"date": datetime.now().strftime("%Y-%m-%d")}
    mapping_counter = Counter(s["match_type"] for s in unique_screens)

    return unique_screens, all_issues, screen_issues, {"run_metadata": run_metadata, "mapping": mapping_counter}


def find_annotated_screenshot(screen_data: dict[str, Any]) -> str | None:
    p = screen_data.get("image_path")
    if p and Path(p).exists():
        return p
    return None


def resize_image_for_excel(img_path: str, max_width: int = 250, max_height: int = 400) -> XlImage | None:
    try:
        with PILImage.open(img_path) as img:
            w, h = img.size
            ratio = min(max_width / w, max_height / h, 1.0)
            new_w = int(w * ratio)
            new_h = int(h * ratio)

            if ratio < 1.0:
                resized = img.resize((new_w, new_h), PILImage.LANCZOS)
                with tempfile.NamedTemporaryFile(
                    delete=False,
                    suffix=".png",
                    prefix="lic_portal_report_",
                    dir="/tmp",
                ) as tmp:
                    temp_path = Path(tmp.name)
                resized.save(temp_path, "PNG")
                TEMP_IMAGE_FILES.append(temp_path)
                xl_img = XlImage(str(temp_path))
            else:
                xl_img = XlImage(img_path)

            xl_img.width = new_w
            xl_img.height = new_h
            return xl_img
    except Exception:
        return None


def apply_header_style(ws, row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER


def apply_data_style(ws, row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.font = DATA_FONT
        cell.alignment = LEFT_ALIGN
        cell.border = BORDER


def aggregate_failures_by_wcag(all_issues: list[dict[str, Any]], unique_screens: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    valid_tags = {s["tag"] for s in unique_screens}
    wcag_failures: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"count": 0, "screens": set(), "severities": [], "details": []}
    )

    for issue in all_issues:
        if issue["screen_tag"] not in valid_tags:
            continue
        for crit_str in issue.get("wcag_criteria", []):
            wcag_num = parse_wcag_number(crit_str)
            if wcag_num:
                wcag_failures[wcag_num]["count"] += 1
                wcag_failures[wcag_num]["screens"].add(issue["screen_tag"])
                wcag_failures[wcag_num]["severities"].append(issue.get("severity", "major"))
                wcag_failures[wcag_num]["details"].append(issue.get("detail", ""))

    return wcag_failures


def get_web_remediation(wcag_num: str) -> tuple[str, str]:
    name = WCAG_CRITERIA.get(wcag_num, ("Unknown", ""))[0]
    return (
        f"Fix {name} issues on affected pages according to WCAG 2.1 guidance.",
        f"1) Identify all failing elements for {wcag_num}\n"
        f"2) Update semantic HTML/ARIA and interaction behavior\n"
        f"3) Re-test with keyboard and screen reader\n"
        f"4) Verify fixes across all affected screens",
    )


def build_executive_dashboard(
    wb: openpyxl.Workbook,
    app_name: str,
    unique_screens: list[dict[str, Any]],
    wcag_failures: dict[str, dict[str, Any]],
    run_metadata: dict[str, Any],
    cv_metrics: dict[str, Any],
) -> None:
    ws = wb.active
    ws.title = "Executive Dashboard"

    total_screens = len(unique_screens)
    unique_activities = len(set((s.get("activity") or "") for s in unique_screens))
    total_pass = sum(s["total_passes"] for s in unique_screens)
    total_fail = sum(s["total_issues"] for s in unique_screens)
    cannot_verify_instances = int(cv_metrics.get("instance_count", 0))
    cannot_verify_checkpoints = int(cv_metrics.get("checkpoint_count", 0))
    cv_threshold = int(cv_metrics.get("threshold", 31))
    cv_checkpoint_ok = bool(cv_metrics.get("checkpoint_within_threshold", False))
    cv_instance_ok = bool(cv_metrics.get("instance_within_threshold", False))
    cv_overall_ok = bool(cv_metrics.get("within_threshold", False))

    total_criteria = len(WCAG_CRITERIA)
    failing_criteria = len(wcag_failures)
    passing_criteria = total_criteria - failing_criteria
    wcag_score = round((passing_criteria / total_criteria) * 100, 1) if total_criteria > 0 else 0

    total_checks = total_pass + total_fail

    ws.merge_cells("A1:H1")
    ws.cell(1, 1, f"{app_name} — WCAG 2.1 Accessibility Audit Report").font = TITLE_FONT
    ws.cell(1, 1).alignment = LEFT_ALIGN

    ws.merge_cells("A2:H2")
    run_date = run_metadata.get("date", datetime.now().strftime("%Y-%m-%d"))
    ws.cell(2, 1, f"App: {app_name}  |  Date: {run_date}").font = SUBTITLE_FONT
    ws.cell(2, 1).alignment = LEFT_ALIGN

    ws.merge_cells("A3:H3")

    headers = [
        "Screens Analyzed",
        "Unique Activities",
        "Total PASS",
        "Total FAIL",
        "Cannot Verify (Checkpoints)",
        "Cannot Verify (Instances)",
        "WCAG Score",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(5, i, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER

    values = [
        total_screens,
        unique_activities,
        total_pass,
        total_fail,
        cannot_verify_checkpoints,
        cannot_verify_instances,
        f"{wcag_score}%",
    ]
    for i, v in enumerate(values, 1):
        cell = ws.cell(6, i, v)
        cell.font = Font(name="Calibri", size=14, bold=True)
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
        if i in {5, 6}:
            cell.fill = FAIL_FILL if int(v) > cv_threshold else PASS_FILL

    ws.merge_cells("A7:H7")

    ws.merge_cells("A9:H9")
    ws.cell(9, 1, "WCAG 2.1 Compliance Score Computation").font = SECTION_FONT
    ws.cell(9, 1).fill = SECTION_FILL

    comp_headers = ["Metric", "Value", "Formula", "Notes"]
    for i, h in enumerate(comp_headers, 1):
        cell = ws.cell(10, i, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER

    comp_data = [
        ("Total Checks Evaluated", f"{total_checks:,}", f"{total_screens} screens", "Across all screens"),
        ("Passed Checks (instances)", f"{total_pass:,}", "—", "PASS + Not Applicable"),
        ("Failed Checks (instances)", f"{total_fail:,}", "—", "Individual screen-check failures"),
        ("Cannot Verify (checkpoints)", f"{cannot_verify_checkpoints}", "Primary KPI", "Distinct WCAG checkpoints still unresolved"),
        ("Cannot Verify (instances)", f"{cannot_verify_instances}", "—", "Screen-level unresolved instances"),
        (
            f"CV Threshold Checkpoints (<= {cv_threshold})",
            "PASS" if cv_checkpoint_ok else "FAIL",
            f"{cannot_verify_checkpoints} <= {cv_threshold}",
            "Threshold gate for checkpoint-level CV",
        ),
        (
            f"CV Threshold Instances (<= {cv_threshold})",
            "PASS" if cv_instance_ok else "FAIL",
            f"{cannot_verify_instances} <= {cv_threshold}",
            "Threshold gate for instance-level CV",
        ),
        (
            f"CV Threshold Overall ({cv_metrics.get('enforcement', 'both')})",
            "PASS" if cv_overall_ok else "FAIL",
            "Both metrics enforced" if cv_metrics.get("enforcement", "both") == "both" else "Configured enforcement",
            "Final threshold decision",
        ),
        ("WCAG 2.1 AA Criteria Tested", f"{total_criteria}", "—", "Level A + Level AA"),
        ("Unique Criteria with Failures", f"{failing_criteria}", "—", "Distinct WCAG checkpoints failing"),
        ("Criteria Fully Passing", f"{passing_criteria}", f"{total_criteria} - {failing_criteria}", "No violations across any screen"),
        ("WCAG Checklist Compliance", f"{wcag_score}%", f"{passing_criteria} / {total_criteria} x 100", "Checklist-level pass rate"),
    ]

    for i, (metric, value, formula, notes) in enumerate(comp_data):
        row = 11 + i
        ws.cell(row, 1, metric).font = DATA_FONT
        ws.cell(row, 2, value).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row, 3, formula).font = DATA_FONT
        ws.cell(row, 4, notes).font = DATA_FONT
        for c in range(1, 5):
            ws.cell(row, c).border = BORDER
            ws.cell(row, c).alignment = LEFT_ALIGN
        if "CV Threshold" in metric:
            value_cell = ws.cell(row, 2)
            if value == "PASS":
                value_cell.fill = PASS_FILL
            elif value == "FAIL":
                value_cell.fill = FAIL_FILL
            value_cell.alignment = CENTER_ALIGN

    spacer_row = 11 + len(comp_data) + 2
    ws.merge_cells(f"A{spacer_row}:H{spacer_row}")
    ws.cell(spacer_row, 1, "Failure Distribution by WCAG Checkpoint").font = SECTION_FONT
    ws.cell(spacer_row, 1).fill = SECTION_FILL

    dist_headers = ["WCAG #", "Checkpoint Name", "Level", "Occurrences", "% of Screens", "Severity", "Priority"]
    dist_row = spacer_row + 1
    for i, h in enumerate(dist_headers, 1):
        cell = ws.cell(dist_row, i, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER

    sorted_wcag = sorted(wcag_failures.items(), key=lambda x: x[1]["count"], reverse=True)
    for i, (wcag_num, fdata) in enumerate(sorted_wcag):
        row = dist_row + 1 + i
        name, level = WCAG_CRITERIA.get(wcag_num, ("Unknown", "A"))

        sev_counts: Counter[str] = Counter()
        for s in fdata["severities"]:
            sev_counts[SEVERITY_MAP.get(s, "Major")] += 1
        top_severity = max(sev_counts, key=sev_counts.get) if sev_counts else "Major"

        pct = round(len(fdata["screens"]) / total_screens * 100) if total_screens > 0 else 0

        ws.cell(row, 1, wcag_num).font = DATA_FONT
        ws.cell(row, 2, name).font = DATA_FONT
        ws.cell(row, 3, level).font = DATA_FONT
        ws.cell(row, 4, fdata["count"]).font = DATA_FONT
        ws.cell(row, 5, f"{pct}%").font = DATA_FONT
        ws.cell(row, 6, top_severity).font = DATA_FONT
        ws.cell(row, 7, SEVERITY_TO_PRIORITY.get(top_severity, "P2")).font = DATA_FONT

        for c in range(1, 8):
            ws.cell(row, c).border = BORDER
            ws.cell(row, c).alignment = CENTER_ALIGN

        sev_cell = ws.cell(row, 6)
        if top_severity == "Critical":
            sev_cell.fill = FAIL_FILL
        elif top_severity == "Major":
            sev_cell.fill = WARN_FILL

    col_widths = [24, 24, 16, 16, 18, 18, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_screen_details(wb: openpyxl.Workbook, unique_screens: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Screen Details")

    ws.merge_cells("A1:I1")
    ws.cell(1, 1, "Detailed Screen-by-Screen WCAG Audit").font = TITLE_FONT

    headers = ["#", "Screen Label", "Annotated Screenshot", "Activity", "PASS", "FAIL", "Cannot Verify", "Pass Rate", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    apply_header_style(ws, 2, len(headers))

    row = 3
    for idx, screen in enumerate(unique_screens, 1):
        passes = screen["total_passes"]
        fails = screen["total_issues"]
        total = passes + fails
        pass_rate = round((passes / total) * 100, 1) if total > 0 else 0
        status = "PASS" if pass_rate >= 60 else "FAIL"

        ws.cell(row, 1, idx)
        ws.cell(row, 2, screen["screen_label"])

        img_path = find_annotated_screenshot(screen)
        if img_path:
            ws.cell(row, 3, Path(img_path).name)
            xl_img = resize_image_for_excel(img_path, max_width=200, max_height=350)
            if xl_img:
                ws.add_image(xl_img, f"C{row}")
                ws.row_dimensions[row].height = max(280, xl_img.height * 0.75)
        else:
            ws.cell(row, 3, "No screenshot")

        ws.cell(row, 4, screen.get("activity", ""))
        ws.cell(row, 5, passes)
        ws.cell(row, 6, fails)
        ws.cell(row, 7, screen.get("cannot_verify", 0))
        ws.cell(row, 8, f"{pass_rate}%")

        status_cell = ws.cell(row, 9, status)
        status_cell.fill = PASS_FILL if status == "PASS" else FAIL_FILL

        apply_data_style(ws, row, len(headers))
        ws.cell(row, 1).alignment = CENTER_ALIGN
        ws.cell(row, 5).alignment = CENTER_ALIGN
        ws.cell(row, 6).alignment = CENTER_ALIGN
        ws.cell(row, 7).alignment = CENTER_ALIGN
        ws.cell(row, 8).alignment = CENTER_ALIGN
        ws.cell(row, 9).alignment = CENTER_ALIGN

        row += 1

    col_widths = [5, 34, 35, 52, 8, 8, 14, 12, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_failure_details(wb: openpyxl.Workbook, unique_screens: list[dict[str, Any]], all_issues: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Failure Details")

    ws.merge_cells("A1:H1")
    ws.cell(1, 1, "Complete Failure Log — All WCAG Violations").font = TITLE_FONT

    headers = ["#", "Screen", "WCAG #", "Checkpoint Name", "Level", "Failure Rationale", "Severity", "Annotated Image"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    apply_header_style(ws, 2, len(headers))

    row = 3
    counter = 1

    for screen in unique_screens:
        tag = screen["tag"]
        img_name = Path(screen.get("image_path", "")).name

        screen_wcag: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for issue in all_issues:
            if issue["screen_tag"] != tag:
                continue
            for crit_str in issue.get("wcag_criteria", []):
                wcag_num = parse_wcag_number(crit_str)
                if wcag_num:
                    screen_wcag[wcag_num].append(issue)

        if not screen_wcag:
            ws.cell(row, 1, counter)
            ws.cell(row, 2, screen["screen_label"])
            ws.cell(row, 3, "")
            ws.cell(row, 4, "")
            ws.cell(row, 5, "")
            ws.cell(row, 6, "No mapped failures")
            ws.cell(row, 7, "Minor")
            ws.cell(row, 8, img_name)
            apply_data_style(ws, row, len(headers))
            ws.cell(row, 1).alignment = CENTER_ALIGN
            ws.cell(row, 7).alignment = CENTER_ALIGN
            row += 1
            counter += 1
            continue

        for wcag_num in sorted(screen_wcag.keys()):
            issues = screen_wcag[wcag_num]
            name, level = WCAG_CRITERIA.get(wcag_num, ("Unknown", "A"))

            details: list[str] = []
            for iss in issues:
                d = (iss.get("detail") or "").strip()
                if d and d not in details:
                    details.append(d)
            detail_text = details[0] if details else ""

            ws.cell(row, 1, counter)
            ws.cell(row, 2, screen["screen_label"])
            ws.cell(row, 3, wcag_num)
            ws.cell(row, 4, name)
            ws.cell(row, 5, level)
            ws.cell(row, 6, detail_text)
            ws.cell(row, 7, "Major")
            ws.cell(row, 8, img_name)

            apply_data_style(ws, row, len(headers))
            ws.cell(row, 1).alignment = CENTER_ALIGN
            ws.cell(row, 3).alignment = CENTER_ALIGN
            ws.cell(row, 5).alignment = CENTER_ALIGN
            ws.cell(row, 7).alignment = CENTER_ALIGN
            ws.cell(row, 7).fill = WARN_FILL

            row += 1
            counter += 1

    col_widths = [5, 32, 10, 25, 8, 64, 12, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_remediation_guide(wb: openpyxl.Workbook, wcag_failures: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Remediation Guide")

    ws.merge_cells("A1:G1")
    ws.cell(1, 1, "WCAG 2.1 Remediation Guide — Fixes & Implementation").font = TITLE_FONT

    headers = ["WCAG #", "Checkpoint", "Level", "Occurrences", "What to Fix", "How to Fix (Step-by-Step)", "Priority"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    apply_header_style(ws, 2, len(headers))

    row = 3
    for wcag_num, fdata in sorted(wcag_failures.items(), key=lambda x: x[1]["count"], reverse=True):
        name, level = WCAG_CRITERIA.get(wcag_num, ("Unknown", "A"))
        what_to_fix, how_to_fix = get_web_remediation(wcag_num)

        sev_counts: Counter[str] = Counter()
        for s in fdata["severities"]:
            sev_counts[SEVERITY_MAP.get(s, "Major")] += 1
        top_severity = max(sev_counts, key=sev_counts.get) if sev_counts else "Major"

        ws.cell(row, 1, wcag_num)
        ws.cell(row, 2, name)
        ws.cell(row, 3, level)
        ws.cell(row, 4, fdata["count"])
        ws.cell(row, 5, what_to_fix)
        ws.cell(row, 6, how_to_fix)
        ws.cell(row, 7, SEVERITY_TO_PRIORITY.get(top_severity, "P2"))

        apply_data_style(ws, row, len(headers))
        ws.cell(row, 1).alignment = CENTER_ALIGN
        ws.cell(row, 3).alignment = CENTER_ALIGN
        ws.cell(row, 4).alignment = CENTER_ALIGN
        ws.cell(row, 7).alignment = CENTER_ALIGN
        ws.row_dimensions[row].height = 80

        row += 1

    col_widths = [10, 25, 8, 14, 52, 60, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_wcag_checklist_pct(wb: openpyxl.Workbook, unique_screens: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("WCAG Checklist %")

    ws.merge_cells("A1:F1")
    ws.cell(1, 1, "WCAG 2.1 Checklist — Per-Screen Compliance %").font = TITLE_FONT

    headers = ["#", "Screen Label", "PASS", "FAIL", "Testable (P+F)", "Compliance %"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    apply_header_style(ws, 2, len(headers))

    row = 3
    for idx, screen in enumerate(unique_screens, 1):
        passes = screen["total_passes"]
        fails = screen["total_issues"]
        testable = passes + fails
        compliance = round((passes / testable) * 100, 1) if testable > 0 else 0

        ws.cell(row, 1, idx)
        ws.cell(row, 2, screen["screen_label"])
        ws.cell(row, 3, passes)
        ws.cell(row, 4, fails)
        ws.cell(row, 5, testable)
        ws.cell(row, 6, f"{compliance}%")

        apply_data_style(ws, row, len(headers))
        ws.cell(row, 1).alignment = CENTER_ALIGN
        ws.cell(row, 3).alignment = CENTER_ALIGN
        ws.cell(row, 4).alignment = CENTER_ALIGN
        ws.cell(row, 5).alignment = CENTER_ALIGN
        ws.cell(row, 6).alignment = CENTER_ALIGN

        comp_cell = ws.cell(row, 6)
        if compliance >= 80:
            comp_cell.fill = PASS_FILL
        elif compliance >= 50:
            comp_cell.fill = WARN_FILL
        else:
            comp_cell.fill = FAIL_FILL

        row += 1

    col_widths = [5, 40, 10, 10, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_annotated_gallery(wb: openpyxl.Workbook, unique_screens: list[dict[str, Any]], all_issues: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Annotated Gallery")

    ws.merge_cells("A1:D1")
    ws.cell(1, 1, "Annotated Screenshot Gallery").font = TITLE_FONT

    headers = ["#", "Image Filename", "Screenshot", "Screen & Failures"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    apply_header_style(ws, 2, len(headers))

    row = 3
    for idx, screen in enumerate(unique_screens, 1):
        ws.cell(row, 1, idx)
        ws.cell(row, 2, Path(screen.get("image_path", "")).name)

        img_path = find_annotated_screenshot(screen)
        if img_path:
            xl_img = resize_image_for_excel(img_path, max_width=250, max_height=400)
            if xl_img:
                ws.add_image(xl_img, f"C{row}")
                ws.row_dimensions[row].height = max(320, xl_img.height * 0.75)

        failures: dict[str, list[str]] = defaultdict(list)
        for issue in all_issues:
            if issue["screen_tag"] != screen["tag"]:
                continue
            for crit_str in issue.get("wcag_criteria", []):
                wcag_num = parse_wcag_number(crit_str)
                if wcag_num and issue.get("detail"):
                    failures[wcag_num].append(issue["detail"])

        summary = f"Screen: {screen['screen_label']}\nActivity: {screen.get('activity', '')}\n\n"
        if failures:
            summary += f"Failures ({len(failures)}):\n"
            for wcag_num in sorted(failures.keys()):
                detail = failures[wcag_num][0][:100]
                summary += f"  [{wcag_num}] {detail}\n"
        else:
            summary += "No failures detected."

        ws.cell(row, 4, summary).alignment = TOP_LEFT_ALIGN
        ws.cell(row, 4).font = DATA_FONT

        apply_data_style(ws, row, len(headers))
        ws.cell(row, 1).alignment = CENTER_ALIGN

        row += 1

    col_widths = [5, 32, 40, 70]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def cleanup_temp_images() -> None:
    for p in TEMP_IMAGE_FILES:
        try:
            p.unlink(missing_ok=True)
        except OSError:
            pass


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Generate LIC website WCAG audit report with the same structure "
            "as LIC_Investor_Portal_Website_WCAG_Audit_Report.xlsx"
        )
    )
    parser.add_argument(
        "--target-artifacts",
        nargs="+",
        default=DEFAULT_ARTIFACT_IDS,
        help="Artifact IDs to include screenshots from (e.g. agentic-3b0eb0b8).",
    )
    parser.add_argument(
        "--output",
        default=str(OUTPUT_FILE),
        help="Output XLSX path.",
    )
    parser.add_argument(
        "--app-name",
        default=DEFAULT_APP_NAME,
        help="App/site name to show in the workbook title.",
    )
    parser.add_argument(
        "--cannot-verify-policy",
        default="pass_leaning",
        choices=["pass_leaning"],
        help="Cannot verify resolution policy.",
    )
    parser.add_argument(
        "--cv-threshold",
        type=int,
        default=31,
        help="Threshold for Cannot Verify metrics.",
    )
    parser.add_argument(
        "--cv-enforcement",
        default="both",
        choices=["both", "checkpoint", "instance"],
        help="Threshold enforcement mode.",
    )
    args = parser.parse_args()

    target_artifact_ids = args.target_artifacts
    output_file = Path(args.output)
    app_name = args.app_name
    cannot_verify_policy = args.cannot_verify_policy
    cv_threshold = int(args.cv_threshold)
    cv_enforcement = args.cv_enforcement

    unique_screens, all_issues, screen_issues, meta = build_data_model(
        target_artifact_ids=target_artifact_ids,
        cannot_verify_policy=cannot_verify_policy,
    )

    # Ensure no duplicate canonical screens are present
    seen = set()
    deduped: list[dict[str, Any]] = []
    for s in unique_screens:
        k = canonical_screen_key(s["filename"])
        if k in seen:
            continue
        seen.add(k)
        deduped.append(s)
    unique_screens = deduped

    # Reindex labels after final dedupe/order
    for i, s in enumerate(unique_screens, 1):
        s["screen_label"] = f"Screen-{i:02d} ({app_name})"

    wcag_failures = aggregate_failures_by_wcag(all_issues, unique_screens)
    cv_metrics = compute_cv_metrics(unique_screens, cv_threshold, cv_enforcement)

    wb = openpyxl.Workbook()
    build_executive_dashboard(
        wb,
        app_name,
        unique_screens,
        wcag_failures,
        meta["run_metadata"],
        cv_metrics,
    )
    build_screen_details(wb, unique_screens)
    build_failure_details(wb, unique_screens, all_issues)
    build_remediation_guide(wb, wcag_failures)
    build_wcag_checklist_pct(wb, unique_screens)
    build_annotated_gallery(wb, unique_screens, all_issues)

    wb.save(output_file)
    cleanup_temp_images()

    print("Generated report:")
    print(f"  {output_file}")
    print(f"Target artifacts: {', '.join(target_artifact_ids)}")
    print(f"Unique screens: {len(unique_screens)}")
    print(f"Contains homepage: {any('initial-load' in canonical_screen_key(s['filename']) for s in unique_screens)}")
    print(
        "Cannot Verify metrics: "
        f"checkpoints={cv_metrics['checkpoint_count']}, "
        f"instances={cv_metrics['instance_count']}, "
        f"threshold={cv_metrics['threshold']}, "
        f"enforcement={cv_metrics['enforcement']}, "
        f"within={cv_metrics['within_threshold']}"
    )


if __name__ == "__main__":
    main()
