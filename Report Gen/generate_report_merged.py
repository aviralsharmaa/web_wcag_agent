#!/usr/bin/env python3
"""
WCAG Accessibility Audit Report Generator for Android Apps.
Reads checklist_reports from wcag_output and generates an XLSX report.
Supports merging data from multiple test runs (e.g., auth flow from separate run).
"""

import json
import os
import re
import sys
import glob
from collections import defaultdict, OrderedDict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
from PIL import Image as PILImage


# ── WCAG Criteria Reference ──────────────────────────────────────────────────
WCAG_CRITERIA = {
    "1.1.1": ("Non-text Content", "A"),
    "1.2.1": ("Audio-only and Video-only (Prerecorded)", "A"),
    "1.2.2": ("Captions (Prerecorded)", "A"),
    "1.2.3": ("Audio Description or Media Alternative (Prerecorded)", "A"),
    "1.2.5": ("Audio Description (Prerecorded)", "AA"),
    "1.3.1": ("Info and Relationships", "A"),
    "1.3.2": ("Meaningful Sequence", "A"),
    "1.3.3": ("Sensory Characteristics", "A"),
    "1.3.4": ("Orientation", "AA"),
    "1.3.5": ("Identify Input Purpose", "AA"),
    "1.4.1": ("Use of Color", "A"),
    "1.4.2": ("Audio Control", "A"),
    "1.4.3": ("Contrast (Minimum)", "AA"),
    "1.4.4": ("Resize Text", "AA"),
    "1.4.5": ("Images of Text", "AA"),
    "1.4.10": ("Reflow", "AA"),
    "1.4.11": ("Non-text Contrast", "AA"),
    "1.4.12": ("Text Spacing", "AA"),
    "1.4.13": ("Content on Hover or Focus", "AA"),
    "2.1.1": ("Keyboard", "A"),
    "2.1.2": ("No Keyboard Trap", "A"),
    "2.1.4": ("Character Key Shortcuts", "A"),
    "2.2.1": ("Timing Adjustable", "A"),
    "2.2.2": ("Pause, Stop, Hide", "A"),
    "2.3.1": ("Three Flashes or Below Threshold", "A"),
    "2.4.1": ("Bypass Blocks", "A"),
    "2.4.2": ("Page Titled", "A"),
    "2.4.3": ("Focus Order", "A"),
    "2.4.5": ("Multiple Ways", "AA"),
    "2.4.6": ("Headings and Labels", "AA"),
    "2.4.7": ("Focus Visible", "AA"),
    "2.5.1": ("Pointer Gestures", "A"),
    "2.5.2": ("Pointer Cancellation", "A"),
    "2.5.3": ("Label in Name", "A"),
    "2.5.4": ("Motion Actuation", "A"),
    "2.5.7": ("Dragging Movements", "AA"),
    "2.5.8": ("Target Size (Minimum)", "AA"),
    "3.1.1": ("Language of Page", "A"),
    "3.1.2": ("Language of Parts", "AA"),
    "3.2.1": ("On Focus", "A"),
    "3.2.2": ("On Input", "A"),
    "3.2.3": ("Consistent Navigation", "AA"),
    "3.2.4": ("Consistent Identification", "AA"),
    "3.2.6": ("Consistent Help", "A"),
    "3.3.1": ("Error Identification", "A"),
    "3.3.2": ("Labels or Instructions", "A"),
    "3.3.3": ("Error Suggestion", "AA"),
    "3.3.4": ("Error Prevention (Legal, Financial, Data)", "AA"),
    "3.3.7": ("Redundant Entry", "A"),
    "3.3.8": ("Accessible Authentication (Minimum)", "AA"),
    "4.1.2": ("Name, Role, Value", "A"),
    "4.1.3": ("Status Messages", "AA"),
    "2.4.11": ("Focus Not Obscured (Minimum)", "AA"),
    "2.4.4": ("Link Purpose (In Context)", "A"),
}

# Map issue_type to WCAG criteria numbers
ISSUE_TYPE_TO_WCAG = {
    "missing_accessible_name": ["4.1.2", "1.1.1"],
    "image_missing_description": ["1.1.1"],
    "small_touch_target": ["2.5.8"],
    "missing_label": ["3.3.2", "1.3.1"],
    "low_contrast": ["1.4.3"],
    "missing_content_description": ["1.1.1"],
    "no_focus_indicator": ["2.4.7"],
    "keyboard_trap": ["2.1.2"],
    "missing_heading_structure": ["1.3.1", "2.4.6"],
    "missing_role": ["4.1.2"],
    "orientation_locked": ["1.3.4"],
    "missing_error_identification": ["3.3.1"],
    "missing_error_suggestion": ["3.3.3"],
    "missing_language": ["3.1.1"],
    "inconsistent_navigation": ["3.2.3"],
    "missing_bypass": ["2.4.1"],
    "missing_page_title": ["2.4.2"],
    "missing_focus_order": ["2.4.3"],
    "sensory_only": ["1.3.3"],
    "color_only": ["1.4.1"],
}

# Severity mapping
SEVERITY_MAP = {
    "high": "Critical",
    "critical": "Critical",
    "medium": "Major",
    "major": "Major",
    "low": "Moderate",
    "moderate": "Moderate",
    "info": "Minor",
    "minor": "Minor",
}

# Priority mapping
SEVERITY_TO_PRIORITY = {
    "Critical": "P1",
    "Major": "P2",
    "Moderate": "P3",
    "Minor": "P4",
}

# ── Styling Constants ────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", size=11, color="555555")
SECTION_FONT = Font(name="Calibri", size=13, bold=True, color="1F4E79")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOP_LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


def parse_wcag_number(criteria_str):
    """Extract WCAG number like '1.1.1' from a string like 'SC 1.1.1 Non-text Content (Level A)'."""
    m = re.search(r'(\d+\.\d+\.\d+)', criteria_str)
    return m.group(1) if m else None


def load_checklist_data_merged(main_dir, auth_dir=None):
    """
    Load all checklist reports and aggregate data.
    Uses auth_dir for the 01_accessible_authentication_minimum checklist if provided.
    """
    all_sources = []  # list of (base_dir, checklist_subdir_name)

    # Determine which checklists come from which directory
    main_reports_dir = os.path.join(main_dir, "checklist_reports")
    if not os.path.isdir(main_reports_dir):
        print(f"ERROR: checklist_reports directory not found at {main_reports_dir}")
        sys.exit(1)

    for cd in sorted(os.listdir(main_reports_dir)):
        cd_path = os.path.join(main_reports_dir, cd)
        if not os.path.isdir(cd_path):
            continue
        # If this is the auth checklist and we have an auth_dir, use auth_dir instead
        if auth_dir and "accessible_authentication" in cd:
            auth_reports_dir = os.path.join(auth_dir, "checklist_reports")
            auth_cd_path = os.path.join(auth_reports_dir, cd)
            if os.path.isdir(auth_cd_path):
                all_sources.append((auth_dir, cd))
                print(f"  Using AUTH dir for: {cd}")
                continue
        all_sources.append((main_dir, cd))

    checklists = []
    all_screens = OrderedDict()  # tag -> screen data
    all_issues = []  # flat list of all issues
    screen_issues = defaultdict(list)  # screen_tag -> [issues]
    screen_passes = defaultdict(int)
    # Track which base_dir each screen's screenshots come from
    screen_base_dirs = {}  # tag -> base_dir

    for base_dir, cd in all_sources:
        cd_path = os.path.join(base_dir, "checklist_reports", cd)

        json_files = glob.glob(os.path.join(cd_path, "checklist_*.json"))
        if not json_files:
            continue

        with open(json_files[0]) as f:
            data = json.load(f)

        cl = data["checklist"]
        checklists.append(cl)

        # Process screen analyses
        for screen in cl.get("screen_analyses", []):
            tag = screen["tag"]
            state_str = screen.get("state_str", tag)
            if tag not in all_screens:
                all_screens[tag] = {
                    "tag": tag,
                    "state_str": state_str,
                    "activity": screen.get("activity", ""),
                    "screen_title": screen.get("screen_title", "Unknown"),
                    "total_issues": 0,
                    "total_passes": 0,
                    "screenshot": screen.get("screenshot", ""),
                    "checklist_dir": cd,
                    "_base_dir": base_dir,
                }
                screen_base_dirs[tag] = base_dir
            all_screens[tag]["total_issues"] += screen["issues_found"]
            all_screens[tag]["total_passes"] += screen["passes_found"]
            screen_passes[tag] += screen["passes_found"]

        # Load issues from issue_index.json
        issue_index_path = os.path.join(cd_path, "issues", "issue_index.json")
        if os.path.exists(issue_index_path):
            with open(issue_index_path) as f:
                issue_data = json.load(f)
            for issue in issue_data.get("issues", []):
                issue["_checklist_name"] = cl["name"]
                issue["_checklist_dir"] = cd
                issue["_base_dir"] = base_dir
                all_issues.append(issue)
                screen_issues[issue["screen_tag"]].append(issue)

    return checklists, all_screens, all_issues, screen_issues, screen_passes, screen_base_dirs


def deduplicate_screens(all_screens, screen_issues, base_dir):
    """
    Deduplicate screens by state_str hash (visual fingerprint).
    Keep the screen tag with the most issues for each unique state hash.
    Returns a list of unique screen records with aggregated data.
    """
    hash_groups = defaultdict(list)  # state_str -> [tags]
    for tag, screen in all_screens.items():
        state_str = screen.get("state_str", tag)
        hash_groups[state_str].append(tag)

    # For each hash group, pick the tag with most issues and aggregate
    unique_screens = []
    for state_str, tags in hash_groups.items():
        # Pick best representative tag (most issues)
        best_tag = max(tags, key=lambda t: all_screens[t]["total_issues"])
        screen = dict(all_screens[best_tag])

        # Aggregate issues and passes across all tags with same hash
        screen["total_issues"] = sum(all_screens[t]["total_issues"] for t in tags)
        screen["total_passes"] = sum(all_screens[t]["total_passes"] for t in tags)
        screen["_all_tags"] = tags  # Keep for issue lookup

        unique_screens.append(screen)

    # Sort by tag (chronological)
    unique_screens.sort(key=lambda s: s["tag"])

    # Generate better screen labels (sequential + title hint)
    title_counts = defaultdict(int)
    for screen in unique_screens:
        base_title = screen["screen_title"]
        title_counts[base_title] += 1
        idx = title_counts[base_title]
        screen["screen_label"] = f"Screen-{unique_screens.index(screen) + 1:02d}"
        if "YouTube" in base_title:
            screen["screen_label"] += f" ({base_title})"
        else:
            screen["screen_label"] += f" (iLearn App)"

    return unique_screens


def find_annotated_screenshot(base_dir, screen_tag, screen_data):
    """Find the annotated screenshot for a given screen tag."""
    # First check the screen's own base_dir
    search_dirs = [screen_data.get("_base_dir", base_dir), base_dir]
    # Remove duplicates while preserving order
    seen = set()
    unique_dirs = []
    for d in search_dirs:
        if d not in seen:
            seen.add(d)
            unique_dirs.append(d)

    for search_dir in unique_dirs:
        checklist_reports_dir = os.path.join(search_dir, "checklist_reports")
        if not os.path.isdir(checklist_reports_dir):
            continue

        # Search across all checklist dirs for annotated version
        for cd in sorted(os.listdir(checklist_reports_dir)):
            annotated_path = os.path.join(
                checklist_reports_dir, cd, "states", "annotated",
                f"state_{screen_tag}_annotated.png"
            )
            if os.path.exists(annotated_path):
                return annotated_path

        # Fallback to raw screenshot in states/
        raw_path = os.path.join(search_dir, "states", f"screen_{screen_tag}.png")
        if os.path.exists(raw_path):
            return raw_path

    return None


def resize_image_for_excel(img_path, max_width=250, max_height=400):
    """Resize image to fit in Excel cell and return XlImage."""
    try:
        with PILImage.open(img_path) as img:
            w, h = img.size
            ratio = min(max_width / w, max_height / h, 1.0)
            new_w = int(w * ratio)
            new_h = int(h * ratio)

            if ratio < 1.0:
                resized = img.resize((new_w, new_h), PILImage.LANCZOS)
                temp_path = img_path + "_resized.png"
                resized.save(temp_path, "PNG")
                xl_img = XlImage(temp_path)
            else:
                xl_img = XlImage(img_path)

            xl_img.width = new_w
            xl_img.height = new_h
            return xl_img
    except Exception as e:
        print(f"  Warning: Could not process image {img_path}: {e}")
        return None


def apply_header_style(ws, row, max_col):
    """Apply header styling to a row."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER


def apply_data_style(ws, row, max_col):
    """Apply data styling to a row."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.font = DATA_FONT
        cell.alignment = LEFT_ALIGN
        cell.border = BORDER


def get_all_tags_set(unique_screens):
    """Get set of all tags (including duplicates) covered by unique screens."""
    tags = set()
    for s in unique_screens:
        tags.update(s.get("_all_tags", [s["tag"]]))
    return tags


def aggregate_failures_by_wcag(all_issues, unique_screens):
    """Aggregate failures by WCAG criterion number across all unique screens."""
    valid_tags = get_all_tags_set(unique_screens)
    wcag_failures = defaultdict(lambda: {"count": 0, "screens": set(), "severities": [], "details": []})

    for issue in all_issues:
        if issue["screen_tag"] not in valid_tags:
            continue
        for crit_str in issue.get("wcag_criteria", []):
            wcag_num = parse_wcag_number(crit_str)
            if wcag_num:
                wcag_failures[wcag_num]["count"] += 1
                wcag_failures[wcag_num]["screens"].add(issue["screen_tag"])
                wcag_failures[wcag_num]["severities"].append(issue.get("severity", "medium"))
                wcag_failures[wcag_num]["details"].append(issue.get("detail", ""))

    return wcag_failures


def get_android_remediation(wcag_num):
    """Return Android-specific remediation guidance for a WCAG criterion."""
    remediations = {
        "1.1.1": (
            "Add contentDescription to all ImageView, ImageButton, and decorative elements. "
            "For decorative images use android:importantForAccessibility=\"no\".",
            "1) Add android:contentDescription to all <ImageView> and <ImageButton> elements\n"
            "2) For decorative images: set android:importantForAccessibility=\"no\"\n"
            "3) For custom views: override getContentDescription() or set via setContentDescription()\n"
            "4) Dynamic images: call setContentDescription() when content changes"
        ),
        "1.3.1": (
            "Use proper semantic structure: headings via android:accessibilityHeading, "
            "group related elements with android:screenReaderFocusable.",
            "1) Set android:accessibilityHeading=\"true\" on heading TextViews\n"
            "2) Group related form fields using a parent ViewGroup\n"
            "3) Use android:labelFor to associate labels with EditText fields\n"
            "4) Ensure list items use proper adapter patterns"
        ),
        "1.3.3": (
            "Do not rely solely on visual position, shape, or color to convey meaning. "
            "Add text labels alongside visual cues.",
            "1) Add text labels to color-coded status indicators\n"
            "2) Don't use 'tap the icon on the left' — name the control\n"
            "3) Add contentDescription that includes meaning, not just visual description\n"
            "4) Pair icons with text labels"
        ),
        "1.4.3": (
            "Ensure text contrast ratio is at least 4.5:1 for normal text and 3:1 for large text (18sp+). "
            "Check all theme colors.",
            "1) Audit all text color / background color pairs in themes and styles\n"
            "2) Use Android Accessibility Scanner to detect contrast issues\n"
            "3) Update colors.xml with accessible color values\n"
            "4) Check disabled state colors still meet minimum ratios"
        ),
        "2.1.1": (
            "Ensure all interactive elements are focusable and operable via external keyboard "
            "and D-pad/switch access.",
            "1) Set android:focusable=\"true\" on custom interactive views\n"
            "2) Handle KeyEvent for Enter/Space on custom views\n"
            "3) Test with TalkBack + keyboard navigation\n"
            "4) Ensure onClickListener views also handle accessibility actions"
        ),
        "2.1.2": (
            "Ensure focus can move away from all components. Modals/dialogs must allow dismissal. "
            "No focus traps in custom views.",
            "1) Test TalkBack navigation — ensure swipe moves through all elements\n"
            "2) Dialogs: ensure dismiss button is focusable\n"
            "3) Custom modal views: manage focus with requestFocus() and clearFocus()\n"
            "4) WebView: ensure JavaScript doesn't trap keyboard focus"
        ),
        "2.4.7": (
            "Ensure focus indicators are visible on all interactive elements when navigated via "
            "keyboard/switch access.",
            "1) Don't set android:background overrides that remove focus state\n"
            "2) Define state_focused drawable states in selectors\n"
            "3) Test with 'Show layout bounds' developer option\n"
            "4) Custom views: implement onFocusChanged() with visual feedback"
        ),
        "2.5.8": (
            "Ensure touch targets are at least 24x24 dp (recommended 48x48 dp). "
            "Add padding to small interactive elements.",
            "1) Set minimum size: android:minWidth=\"48dp\" android:minHeight=\"48dp\"\n"
            "2) Use TouchDelegate to expand touch area without visual change\n"
            "3) Add padding to small icon buttons\n"
            "4) Check RecyclerView item touch targets"
        ),
        "3.3.2": (
            "Provide visible labels for all form inputs. Use android:hint and android:labelFor "
            "to associate labels programmatically.",
            "1) Add android:hint to all EditText fields\n"
            "2) Add visible TextView labels with android:labelFor pointing to the input\n"
            "3) Use TextInputLayout with hint for Material Design inputs\n"
            "4) Ensure auto-fill hints are set via android:autofillHints"
        ),
        "4.1.2": (
            "Ensure all interactive elements have accessible name, role, and state. "
            "Use contentDescription, roleDescription, and state descriptions.",
            "1) Add android:contentDescription to ImageButtons and icon-only controls\n"
            "2) Set AccessibilityNodeInfo role via AccessibilityDelegate\n"
            "3) Update stateDescription for toggles/checkboxes\n"
            "4) Custom views: implement onInitializeAccessibilityNodeInfo()"
        ),
        "4.1.3": (
            "Ensure status messages are announced by screen readers without receiving focus. "
            "Use announceForAccessibility() or live regions.",
            "1) Use android:accessibilityLiveRegion=\"polite\" for status areas\n"
            "2) Call view.announceForAccessibility() for transient messages\n"
            "3) Snackbar/Toast: ensure TalkBack announces content\n"
            "4) Loading indicators: announce start/completion states"
        ),
    }

    if wcag_num in remediations:
        return remediations[wcag_num]

    # Generic fallback
    name = WCAG_CRITERIA.get(wcag_num, ("Unknown", ""))[0]
    return (
        f"Review and fix {name} violations as per WCAG 2.1 guidelines.",
        f"1) Identify all elements violating {wcag_num}\n"
        f"2) Apply Android accessibility best practices\n"
        f"3) Test with TalkBack and Accessibility Scanner\n"
        f"4) Verify fix across all affected screens"
    )


# ── Sheet Builders ───────────────────────────────────────────────────────────

def build_executive_dashboard(wb, app_name, unique_screens, all_issues, wcag_failures, checklists, run_metadata):
    """Build the Executive Dashboard sheet."""
    ws = wb.active
    ws.title = "Executive Dashboard"

    total_screens = len(unique_screens)
    unique_activities = len(set(s["activity"] for s in unique_screens))
    total_pass = sum(s["total_passes"] for s in unique_screens)
    total_fail = sum(s["total_issues"] for s in unique_screens)

    # Count WCAG criteria tested and failing
    total_criteria = len(WCAG_CRITERIA)
    failing_criteria = len(wcag_failures)
    passing_criteria = total_criteria - failing_criteria
    wcag_score = round((passing_criteria / total_criteria) * 100, 1) if total_criteria > 0 else 0

    cannot_verify = 0
    total_checks = total_pass + total_fail

    # Row 1: Title
    ws.merge_cells("A1:H1")
    ws.cell(1, 1, f"{app_name} — WCAG 2.1 Accessibility Audit Report").font = TITLE_FONT
    ws.cell(1, 1).alignment = LEFT_ALIGN

    # Row 2: Subtitle
    ws.merge_cells("A2:H2")
    run_date = run_metadata.get("date", datetime.now().strftime("%B %Y"))
    ws.cell(2, 1, f"App: {app_name}  |  Date: {run_date}").font = SUBTITLE_FONT
    ws.cell(2, 1).alignment = LEFT_ALIGN

    # Row 3: spacer
    ws.merge_cells("A3:H3")

    # Row 5: Summary headers
    headers = ["Screens Analyzed", "Unique Activities", "Total PASS", "Total FAIL",
               "Cannot Verify", "WCAG Score"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(5, i, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER

    # Row 6: Summary values
    values = [total_screens, unique_activities, total_pass, total_fail, cannot_verify, f"{wcag_score}%"]
    for i, v in enumerate(values, 1):
        cell = ws.cell(6, i, v)
        cell.font = Font(name="Calibri", size=14, bold=True)
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER

    # Row 7: spacer
    ws.merge_cells("A7:H7")

    # Row 9: Score computation section
    ws.merge_cells("A9:H9")
    ws.cell(9, 1, "WCAG 2.1 Compliance Score Computation").font = SECTION_FONT
    ws.cell(9, 1).fill = SECTION_FILL

    # Score computation table
    comp_headers = ["Metric", "Value", "Formula", "Notes"]
    for i, h in enumerate(comp_headers, 1):
        cell = ws.cell(10, i, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER

    comp_data = [
        ("Total Checks Evaluated", f"{total_checks:,}", f"{total_screens} screens", "Across all screens"),
        ("Passed Checks (instances)", f"{total_pass:,}", "—", "Individual screen-check passes"),
        ("Failed Checks (instances)", f"{total_fail:,}", "—", "Individual screen-check failures"),
        ("Cannot Verify (instances)", f"{cannot_verify}", "—", "Requires manual/AT testing"),
        ("WCAG 2.1 AA Criteria Tested", f"{total_criteria}", "—",
         f"{sum(1 for v in WCAG_CRITERIA.values() if v[1] == 'A')} Level A + "
         f"{sum(1 for v in WCAG_CRITERIA.values() if v[1] == 'AA')} Level AA"),
        ("Unique Criteria with Failures", f"{failing_criteria}", "—", "Distinct WCAG checkpoints failing"),
        ("Criteria Fully Passing", f"{passing_criteria}", f"{total_criteria} - {failing_criteria}",
         "No violations across any screen"),
        ("WCAG Checklist Compliance", f"{wcag_score}%", f"{passing_criteria} / {total_criteria} x 100",
         "Checklist-level pass rate"),
    ]
    for i, (metric, value, formula, notes) in enumerate(comp_data):
        row = 11 + i
        ws.cell(row, 1, metric).font = DATA_FONT
        ws.cell(row, 2, value).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row, 3, formula).font = DATA_FONT
        ws.cell(row, 4, notes).font = DATA_FONT
        for c in range(1, 5):
            ws.cell(row, c).border = BORDER
            ws.cell(row, c).alignment = LEFT_ALIGN

    # Spacer
    spacer_row = 11 + len(comp_data) + 2

    # Failure Distribution section
    ws.merge_cells(f"A{spacer_row}:H{spacer_row}")
    ws.cell(spacer_row, 1, "Failure Distribution by WCAG Checkpoint").font = SECTION_FONT
    ws.cell(spacer_row, 1).fill = SECTION_FILL

    dist_headers = ["WCAG #", "Checkpoint Name", "Level", "Occurrences", "% of Screens", "Severity", "Priority"]
    dist_row = spacer_row + 1
    for i, h in enumerate(dist_headers, 1):
        cell = ws.cell(dist_row, i, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER

    sorted_wcag = sorted(wcag_failures.items(), key=lambda x: x[1]["count"], reverse=True)
    for i, (wcag_num, fdata) in enumerate(sorted_wcag):
        row = dist_row + 1 + i
        name, level = WCAG_CRITERIA.get(wcag_num, ("Unknown", "A"))

        # Determine most common severity
        sev_counts = defaultdict(int)
        for s in fdata["severities"]:
            sev_counts[SEVERITY_MAP.get(s, "Moderate")] += 1
        top_severity = max(sev_counts, key=sev_counts.get) if sev_counts else "Moderate"

        pct = round(len(fdata["screens"]) / total_screens * 100) if total_screens > 0 else 0

        ws.cell(row, 1, wcag_num).font = DATA_FONT
        ws.cell(row, 2, name).font = DATA_FONT
        ws.cell(row, 3, level).font = DATA_FONT
        ws.cell(row, 4, fdata["count"]).font = DATA_FONT
        ws.cell(row, 5, f"{pct}%").font = DATA_FONT
        ws.cell(row, 6, top_severity).font = DATA_FONT
        ws.cell(row, 7, SEVERITY_TO_PRIORITY.get(top_severity, "P3")).font = DATA_FONT

        for c in range(1, 8):
            ws.cell(row, c).border = BORDER
            ws.cell(row, c).alignment = CENTER_ALIGN

        # Color severity
        sev_cell = ws.cell(row, 6)
        if top_severity == "Critical":
            sev_cell.fill = FAIL_FILL
        elif top_severity == "Major":
            sev_cell.fill = WARN_FILL
        elif top_severity == "Moderate":
            sev_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Column widths
    col_widths = [22, 22, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_screen_details(wb, unique_screens, screen_issues, base_dir):
    """Build the Screen Details sheet with annotated screenshots."""
    ws = wb.create_sheet("Screen Details")

    # Title
    ws.merge_cells("A1:I1")
    ws.cell(1, 1, "Detailed Screen-by-Screen WCAG Audit").font = TITLE_FONT

    # Headers
    headers = ["#", "Screen Label", "Annotated Screenshot", "Activity",
               "PASS", "FAIL", "Cannot Verify", "Pass Rate", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    apply_header_style(ws, 2, len(headers))

    row = 3
    for idx, screen in enumerate(unique_screens, 1):
        tag = screen["tag"]
        passes = screen["total_passes"]
        fails = screen["total_issues"]
        total = passes + fails
        pass_rate = round(passes / total * 100, 1) if total > 0 else 0
        status = "PASS" if pass_rate >= 60 else "FAIL"

        ws.cell(row, 1, idx)
        ws.cell(row, 2, screen.get("screen_label", screen["screen_title"]))

        # Find and insert annotated screenshot
        img_path = find_annotated_screenshot(base_dir, tag, screen)
        img_filename = ""
        if img_path:
            img_filename = os.path.basename(img_path)
            ws.cell(row, 3, img_filename)
            xl_img = resize_image_for_excel(img_path, max_width=200, max_height=350)
            if xl_img:
                cell_ref = f"C{row}"
                ws.add_image(xl_img, cell_ref)
                ws.row_dimensions[row].height = max(280, xl_img.height * 0.75)
        else:
            ws.cell(row, 3, "No screenshot")

        ws.cell(row, 4, screen["activity"].split("/")[-1] if "/" in screen["activity"] else screen["activity"])
        ws.cell(row, 5, passes)
        ws.cell(row, 6, fails)
        ws.cell(row, 7, 0)
        ws.cell(row, 8, f"{pass_rate}%")

        status_cell = ws.cell(row, 9, status)
        status_cell.fill = PASS_FILL if status == "PASS" else FAIL_FILL

        apply_data_style(ws, row, len(headers))
        ws.cell(row, 1).alignment = CENTER_ALIGN
        ws.cell(row, 5).alignment = CENTER_ALIGN
        ws.cell(row, 6).alignment = CENTER_ALIGN
        ws.cell(row, 7).alignment = CENTER_ALIGN
        ws.cell(row, 8).alignment = CENTER_ALIGN
        ws.cell(row, 9).alignment = CENTER_ALIGN

        row += 1

    # Column widths
    col_widths = [5, 30, 35, 40, 8, 8, 14, 12, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_failure_details(wb, unique_screens, all_issues, base_dir):
    """Build the Failure Details sheet."""
    ws = wb.create_sheet("Failure Details")

    ws.merge_cells("A1:H1")
    ws.cell(1, 1, "Complete Failure Log — All WCAG Violations").font = TITLE_FONT

    headers = ["#", "Screen", "WCAG #", "Checkpoint Name", "Level",
               "Failure Rationale", "Severity", "Annotated Image"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    apply_header_style(ws, 2, len(headers))

    row = 3
    counter = 1
    seen_failures = set()

    for screen in unique_screens:
        tag = screen["tag"]
        screen_tags = set(screen.get("_all_tags", [tag]))
        screen_title = screen.get("screen_label", screen["screen_title"])

        # Collect unique WCAG failures for this screen (across all equivalent tags)
        screen_wcag_issues = defaultdict(list)
        for issue in all_issues:
            if issue["screen_tag"] not in screen_tags:
                continue
            for crit_str in issue.get("wcag_criteria", []):
                wcag_num = parse_wcag_number(crit_str)
                if wcag_num:
                    key = (tag, wcag_num)
                    if key not in seen_failures:
                        screen_wcag_issues[wcag_num].append(issue)
                        seen_failures.add(key)

        # Find annotated screenshot for this screen
        img_path = find_annotated_screenshot(base_dir, tag, screen)
        img_filename = os.path.basename(img_path) if img_path else ""

        for wcag_num in sorted(screen_wcag_issues.keys()):
            issues_for_criteria = screen_wcag_issues[wcag_num]
            name, level = WCAG_CRITERIA.get(wcag_num, ("Unknown", "A"))

            # Combine details
            details = []
            severities = []
            for iss in issues_for_criteria:
                if iss.get("detail") and iss["detail"] not in details:
                    details.append(iss["detail"])
                severities.append(iss.get("severity", "medium"))

            detail_text = details[0] if len(details) == 1 else "; ".join(details[:3])
            if len(details) > 3:
                detail_text += f" (+{len(details) - 3} more)"

            top_sev = SEVERITY_MAP.get(
                max(set(severities), key=severities.count) if severities else "medium",
                "Moderate"
            )

            ws.cell(row, 1, counter)
            ws.cell(row, 2, screen_title)
            ws.cell(row, 3, wcag_num)
            ws.cell(row, 4, name)
            ws.cell(row, 5, level)
            ws.cell(row, 6, detail_text)
            ws.cell(row, 7, top_sev)
            ws.cell(row, 8, img_filename)

            apply_data_style(ws, row, len(headers))
            ws.cell(row, 1).alignment = CENTER_ALIGN
            ws.cell(row, 5).alignment = CENTER_ALIGN
            ws.cell(row, 7).alignment = CENTER_ALIGN

            # Color severity
            sev_cell = ws.cell(row, 7)
            if top_sev == "Critical":
                sev_cell.fill = FAIL_FILL
            elif top_sev == "Major":
                sev_cell.fill = WARN_FILL

            row += 1
            counter += 1

    # Column widths
    col_widths = [5, 25, 10, 25, 8, 60, 12, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_remediation_guide(wb, wcag_failures):
    """Build the Remediation Guide sheet."""
    ws = wb.create_sheet("Remediation Guide")

    ws.merge_cells("A1:G1")
    ws.cell(1, 1, "WCAG 2.1 Remediation Guide — Fixes & Implementation").font = TITLE_FONT

    headers = ["WCAG #", "Checkpoint", "Level", "Occurrences", "What to Fix",
               "How to Fix (Step-by-Step)", "Priority"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    apply_header_style(ws, 2, len(headers))

    sorted_wcag = sorted(wcag_failures.items(), key=lambda x: x[1]["count"], reverse=True)
    row = 3
    for wcag_num, fdata in sorted_wcag:
        name, level = WCAG_CRITERIA.get(wcag_num, ("Unknown", "A"))
        what_to_fix, how_to_fix = get_android_remediation(wcag_num)

        sev_counts = defaultdict(int)
        for s in fdata["severities"]:
            sev_counts[SEVERITY_MAP.get(s, "Moderate")] += 1
        top_severity = max(sev_counts, key=sev_counts.get) if sev_counts else "Moderate"

        ws.cell(row, 1, wcag_num)
        ws.cell(row, 2, name)
        ws.cell(row, 3, level)
        ws.cell(row, 4, fdata["count"])
        ws.cell(row, 5, what_to_fix)
        ws.cell(row, 6, how_to_fix)
        ws.cell(row, 7, SEVERITY_TO_PRIORITY.get(top_severity, "P3"))

        apply_data_style(ws, row, len(headers))
        ws.cell(row, 1).alignment = CENTER_ALIGN
        ws.cell(row, 3).alignment = CENTER_ALIGN
        ws.cell(row, 4).alignment = CENTER_ALIGN
        ws.cell(row, 7).alignment = CENTER_ALIGN

        # Set row height for wrapped text
        ws.row_dimensions[row].height = 80

        row += 1

    col_widths = [10, 25, 8, 14, 50, 60, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_wcag_checklist_pct(wb, unique_screens, screen_issues):
    """Build the WCAG Checklist % sheet."""
    ws = wb.create_sheet("WCAG Checklist %")

    ws.merge_cells("A1:F1")
    ws.cell(1, 1, "WCAG 2.1 Checklist — Per-Screen Compliance %").font = TITLE_FONT

    headers = ["#", "Screen Label", "PASS", "FAIL", "Testable (P+F)", "Compliance %"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    apply_header_style(ws, 2, len(headers))

    row = 3
    for idx, screen in enumerate(unique_screens, 1):
        passes = screen["total_passes"]
        fails = screen["total_issues"]
        testable = passes + fails
        compliance = round(passes / testable * 100, 1) if testable > 0 else 0

        ws.cell(row, 1, idx)
        ws.cell(row, 2, screen.get("screen_label", screen["screen_title"]))
        ws.cell(row, 3, passes)
        ws.cell(row, 4, fails)
        ws.cell(row, 5, testable)
        ws.cell(row, 6, f"{compliance}%")

        apply_data_style(ws, row, len(headers))
        ws.cell(row, 1).alignment = CENTER_ALIGN
        ws.cell(row, 3).alignment = CENTER_ALIGN
        ws.cell(row, 4).alignment = CENTER_ALIGN
        ws.cell(row, 5).alignment = CENTER_ALIGN
        ws.cell(row, 6).alignment = CENTER_ALIGN

        # Color compliance
        comp_cell = ws.cell(row, 6)
        if compliance >= 80:
            comp_cell.fill = PASS_FILL
        elif compliance >= 50:
            comp_cell.fill = WARN_FILL
        else:
            comp_cell.fill = FAIL_FILL

        row += 1

    col_widths = [5, 40, 10, 10, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_annotated_gallery(wb, unique_screens, all_issues, base_dir):
    """Build the Annotated Gallery sheet with screenshots and failure summaries."""
    ws = wb.create_sheet("Annotated Gallery")

    ws.merge_cells("A1:D1")
    ws.cell(1, 1, "Annotated Screenshot Gallery").font = TITLE_FONT

    headers = ["#", "Image Filename", "Screenshot", "Screen & Failures"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    apply_header_style(ws, 2, len(headers))

    row = 3
    for idx, screen in enumerate(unique_screens, 1):
        tag = screen["tag"]
        screen_tags = set(screen.get("_all_tags", [tag]))
        screen_title = screen.get("screen_label", screen["screen_title"])

        img_path = find_annotated_screenshot(base_dir, tag, screen)
        img_filename = os.path.basename(img_path) if img_path else "N/A"

        ws.cell(row, 1, idx)
        ws.cell(row, 2, img_filename)

        # Insert screenshot
        if img_path:
            xl_img = resize_image_for_excel(img_path, max_width=250, max_height=400)
            if xl_img:
                ws.add_image(xl_img, f"C{row}")
                ws.row_dimensions[row].height = max(320, xl_img.height * 0.75)

        # Build failure summary text
        screen_failures = defaultdict(list)
        for issue in all_issues:
            if issue["screen_tag"] not in screen_tags:
                continue
            for crit_str in issue.get("wcag_criteria", []):
                wcag_num = parse_wcag_number(crit_str)
                if wcag_num and issue.get("detail"):
                    screen_failures[wcag_num].append(issue["detail"])

        # Deduplicate
        unique_failures = {}
        for wcag_num, details in screen_failures.items():
            unique_details = list(dict.fromkeys(details))
            unique_failures[wcag_num] = unique_details[0]

        summary = f"Screen: {screen_title}\nActivity: {screen['activity']}\n\n"
        if unique_failures:
            summary += f"Failures ({len(unique_failures)}):\n"
            for wcag_num in sorted(unique_failures.keys()):
                detail = unique_failures[wcag_num]
                summary += f"  [{wcag_num}] {detail[:100]}\n"
        else:
            summary += "No failures detected."

        ws.cell(row, 4, summary).alignment = TOP_LEFT_ALIGN
        ws.cell(row, 4).font = DATA_FONT

        apply_data_style(ws, row, len(headers))
        ws.cell(row, 1).alignment = CENTER_ALIGN

        row += 1

    col_widths = [5, 30, 40, 70]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    # Main data directory (15 checklists)
    main_dir = "/Users/aviral/Projects/android-test/projects/Archive/wcag_output/com-icicidirect-ilearn__v1-0-29__20260317_155341"
    # Auth flow directory (authentication checklist override)
    auth_dir = "/Users/aviral/Projects/android-test/projects/Archive/wcag_output/com-icicidirect-ilearn__v1-0-29__20260317_135348"

    if len(sys.argv) >= 2:
        main_dir = sys.argv[1]
    if len(sys.argv) >= 3:
        auth_dir = sys.argv[2]

    if not os.path.isdir(main_dir):
        print(f"ERROR: Main directory not found: {main_dir}")
        sys.exit(1)

    print(f"Loading data from: {main_dir}")
    if auth_dir and os.path.isdir(auth_dir):
        print(f"Auth flow override from: {auth_dir}")
    else:
        auth_dir = None

    checklists, all_screens, all_issues, screen_issues, screen_passes, screen_base_dirs = \
        load_checklist_data_merged(main_dir, auth_dir)

    print(f"  Total checklists: {len(checklists)}")
    print(f"  Total screens (raw): {len(all_screens)}")
    print(f"  Total issues: {len(all_issues)}")

    # Deduplicate screens
    unique_screens = deduplicate_screens(all_screens, screen_issues, main_dir)
    print(f"  Unique screens (deduplicated): {len(unique_screens)}")

    # Aggregate WCAG failures
    wcag_failures = aggregate_failures_by_wcag(all_issues, unique_screens)
    print(f"  Unique WCAG criteria with failures: {len(wcag_failures)}")

    # Load run metadata
    run_metadata = {"date": "March 2026"}
    metadata_path = os.path.join(main_dir, "run_metadata.json")
    if os.path.exists(metadata_path):
        with open(metadata_path) as f:
            meta = json.load(f)
        run_metadata["date"] = meta.get("started_at_utc", "March 2026")[:10]

    # Determine app name
    app_name = "ICICI Direct iLearn"

    print(f"\nGenerating report for: {app_name}")

    # Create workbook
    wb = openpyxl.Workbook()

    print("  Building Executive Dashboard...")
    build_executive_dashboard(wb, app_name, unique_screens, all_issues, wcag_failures, checklists, run_metadata)

    print("  Building Screen Details...")
    build_screen_details(wb, unique_screens, screen_issues, main_dir)

    print("  Building Failure Details...")
    build_failure_details(wb, unique_screens, all_issues, main_dir)

    print("  Building Remediation Guide...")
    build_remediation_guide(wb, wcag_failures)

    print("  Building WCAG Checklist %...")
    build_wcag_checklist_pct(wb, unique_screens, screen_issues)

    print("  Building Annotated Gallery...")
    build_annotated_gallery(wb, unique_screens, all_issues, main_dir)

    # Save
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_file = os.path.join(output_dir, f"{app_name.replace(' ', '_')}_WCAG_Audit_Report_20260317.xlsx")
    wb.save(output_file)
    print(f"\nReport saved to: {output_file}")
    print("Done!")


if __name__ == "__main__":
    main()
