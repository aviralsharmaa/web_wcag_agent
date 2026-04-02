from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from accessibility_scanner.agent.flow_runner import AgenticFlowRunner
from accessibility_scanner.checklist_registry import load_checklist_specs
from accessibility_scanner.checklist_reports import generate_checklist_reports
from accessibility_scanner.xlsx_report import REPORT_STANDARD, generate_xlsx_report


def _sample_checklist_evaluation() -> dict:
    spec = load_checklist_specs()[0]
    return {
        **spec.to_dict(),
        "screen_label": "01-initial-load",
        "page_url": "https://example.gov/",
        "page_title": "Example",
        "status": "Fail",
        "rationale": "Missing alt text for hero image.",
        "manual_required": True,
        "machine_pass": False,
        "selector_or_target": "#hero-image",
        "issues": [
            {
                "issue_type": "non_text_content",
                "severity": "critical",
                "detail": "Missing alt text for hero image.",
                "bounds": {"x": 10, "y": 10, "width": 50, "height": 40},
                "xml_line": None,
                "selector": "#hero-image",
                "tag": "img",
                "wcag_criteria": ["1.1.1"],
                "source": "wcag_result",
            }
        ],
        "evidence": {
            "screenshot": "artifacts/run/01.png",
            "dom_dump": "artifacts/run/01-dom.html",
            "page_info_dump": "artifacts/run/01-page-info.json",
            "wcag_results_dump": "artifacts/run/01-wcag-results.json",
            "required_evidence_for_llm": spec.required_evidence_for_llm,
        },
        "llm_validation_payload": {
            **spec.to_dict(),
            "screen_label": "01-initial-load",
            "page_url": "https://example.gov/",
            "machine_result": {"status": "Fail", "rationale": "Missing alt text for hero image."},
            "issues": [
                {
                    "issue_type": "non_text_content",
                    "severity": "critical",
                    "detail": "Missing alt text for hero image.",
                    "bounds": {"x": 10, "y": 10, "width": 50, "height": 40},
                    "selector": "#hero-image",
                    "tag": "img",
                    "wcag_criteria": ["1.1.1"],
                    "source": "wcag_result",
                }
            ],
            "evidence": {"screenshot": "artifacts/run/01.png"},
        },
    }


def _sample_screen() -> dict:
    return {
        "label": "01-initial-load",
        "url": "https://example.gov/",
        "unique_key": "https://example.gov|abc123|def456",
        "action_source": "llm",
        "screenshot": "artifacts/run/01.png",
        "annotated_screenshot": "artifacts/run/01-annotated.png",
        "dom_dump": "artifacts/run/01-dom.html",
        "page_info_dump": "artifacts/run/01-page-info.json",
        "wcag_results_dump": "artifacts/run/01-wcag-results.json",
        "wcag_summary_dump": "artifacts/run/01-wcag-summary.json",
        "annotation_metadata": {
            "strategy": "one-screenshot-one-annotation",
            "checkpoint": "1.4.3",
            "mode": "bbox",
            "map_quality": "exact",
            "selector": "p.bad",
            "target_tag": "p",
        },
        "scroll_screenshots": ["artifacts/run/01-fold2.png"],
        "scroll_annotated_screenshots": ["artifacts/run/01-fold2-annotated.png"],
        "wcag_summary": {
            "total_checks": 56,
            "pass": 20,
            "fail": 1,
            "cannot_verify": 30,
            "not_applicable": 5,
            "failures": [
                {
                    "checkpoint": "1.4.3",
                    "rationale": "Low contrast detected.",
                    "page": "https://example.gov/",
                }
            ],
        },
        "wcag_results": [
            {
                "checkpoint_id": "1.4.3",
                "status": "Fail",
                "rationale": "Low contrast detected.",
            },
            {
                "checkpoint_id": "1.4.11",
                "status": "Pass",
                "rationale": "Non-text contrast looks sufficient.",
            },
            {
                "checkpoint_id": "4.1.1",
                "status": "Pass",
                "rationale": "No parsing mismatches.",
            },
        ],
        "contrast_evidence": {
            "text_sampled": 8,
            "non_text_sampled": 4,
            "text_failures": [{"selector": "p.bad", "ratio": 4.1, "required_ratio": 4.5, "pass": False}],
            "non_text_failures": [],
        },
        "checklist_evaluations": [_sample_checklist_evaluation()],
    }


def test_agentic_report_contains_full_screen_payload() -> None:
    runner = AgenticFlowRunner(config_path="config/hdfc_sky_login.json", headless=True)
    runner.screen_results = [_sample_screen()]
    runner.action_trace = [{"action": "click", "source": "llm", "url": "https://example.gov/"}]

    report = runner._build_report("testrun01")
    screen = report["screens"][0]

    assert report["standard"] == REPORT_STANDARD
    assert "action_trace" in report
    assert "wcag_results" in screen
    assert "contrast_evidence" in screen
    assert "checklist_evaluations" in screen
    assert report["checklist_rollup"]
    assert screen["action_source"] == "llm"
    assert screen["annotation_metadata"]["checkpoint"] == "1.4.3"
    assert screen["scroll_annotated_screenshots"]
    assert screen["dom_dump"].endswith("-dom.html")
    assert "cannot_verify_metrics" in report
    assert report["cannot_verify_metrics"]["checkpoint_count"] == 0
    assert report["cannot_verify_metrics"]["instance_count"] == 0


def test_evidence_index_links_all_required_files() -> None:
    runner = AgenticFlowRunner(config_path="config/hdfc_sky_login.json", headless=True)
    runner.screen_results = [_sample_screen()]

    index = runner._build_evidence_index("testrun01")
    assert index["screens"]
    evidence = index["screens"][0]["evidence"]
    assert evidence["screenshot"].endswith(".png")
    assert evidence["annotated_screenshot"].endswith("-annotated.png")
    assert evidence["dom_dump"].endswith("-dom.html")
    assert evidence["page_info_dump"].endswith("-page-info.json")
    assert evidence["wcag_results_dump"].endswith("-wcag-results.json")
    assert index["screens"][0]["checklist_evaluations"][0]["sc_id"] == "1.1.1"


def test_xlsx_uses_full_wcag_results_and_standard_label(tmp_path: Path) -> None:
    report = {
        "run_id": "testrun01",
        "config": "Sample Config",
        "standard": REPORT_STANDARD,
        "screens_analyzed": 1,
        "urls_visited": ["https://example.gov/"],
        "totals": {"pass": 20, "fail": 1, "cannot_verify": 30},
        "all_failures": [],
        "cannot_verify_policy": "pass_leaning",
        "cannot_verify_threshold": 31,
        "cannot_verify_enforcement": "both",
        "cannot_verify_metrics": {
            "checkpoint_count": 0,
            "instance_count": 0,
            "threshold": 31,
            "enforcement": "both",
            "checkpoint_within_threshold": True,
            "instance_within_threshold": True,
            "within_threshold": True,
        },
        "checklist_rollup": [
            {
                **_sample_checklist_evaluation(),
                "aggregate_status": "Fail",
                "screen_evaluations": [_sample_checklist_evaluation()],
                "pages": ["https://example.gov/"],
            }
        ],
        "screens": [_sample_screen()],
    }

    xlsx_path = tmp_path / "wcag_report.xlsx"
    generate_xlsx_report(report, str(xlsx_path))
    wb = load_workbook(str(xlsx_path))

    ws_summary = wb["Summary"]
    summary_rows = {ws_summary.cell(row=i, column=1).value: ws_summary.cell(row=i, column=2).value for i in range(2, ws_summary.max_row + 1)}
    assert summary_rows["Standard"] == REPORT_STANDARD
    assert summary_rows["Total CANNOT VERIFY (Checkpoints)"] == 0
    assert summary_rows["CV Threshold Overall"] == "PASS"

    ws_wcag = wb["WCAG Guidelines"]
    rows = {
        ws_wcag.cell(row=i, column=1).value: ws_wcag.cell(row=i, column=5).value
        for i in range(2, ws_wcag.max_row + 1)
    }
    assert rows["1.4.3"] == "Fail"
    assert rows["4.1.1"] == "Pass"
    assert "Checklist Detail" in wb.sheetnames


def test_report_cv_metrics_count_checkpoint_and_instance() -> None:
    runner = AgenticFlowRunner(config_path="config/hdfc_sky_login.json", headless=True)
    screen = _sample_screen()
    screen["wcag_results"] = [
        {"checkpoint_id": "2.4.3", "status": "Cannot verify automatically", "rationale": "Needs manual order check"},
        {"checkpoint_id": "1.4.3", "status": "Fail", "rationale": "Contrast fail"},
    ]
    screen["wcag_summary"]["cannot_verify"] = 1
    screen["wcag_summary"]["fail"] = 1
    runner.screen_results = [screen]

    report = runner._build_report("testrun01")
    cv = report["cannot_verify_metrics"]
    assert cv["instance_count"] == 1
    assert cv["checkpoint_count"] == 1


def test_generate_checklist_reports_creates_organized_folder_structure(tmp_path: Path) -> None:
    screenshot = tmp_path / "01.png"
    Image.new("RGB", (160, 120), (255, 255, 255)).save(screenshot)
    dom_dump = tmp_path / "01-dom.html"
    dom_dump.write_text("<html><body><img id='hero-image'></body></html>", encoding="utf-8")
    page_info = tmp_path / "01-page-info.json"
    page_info.write_text('{"title":"Example"}', encoding="utf-8")
    wcag_dump = tmp_path / "01-wcag-results.json"
    wcag_dump.write_text('[{"checkpoint_id":"1.1.1","status":"Fail"}]', encoding="utf-8")

    runner = AgenticFlowRunner(config_path="config/hdfc_sky_login.json", headless=True)
    screen = _sample_screen()
    screen["screenshot"] = str(screenshot)
    screen["dom_dump"] = str(dom_dump)
    screen["page_info_dump"] = str(page_info)
    screen["wcag_results_dump"] = str(wcag_dump)
    screen["checklist_evaluations"][0]["evidence"]["screenshot"] = str(screenshot)
    screen["checklist_evaluations"][0]["evidence"]["dom_dump"] = str(dom_dump)
    screen["checklist_evaluations"][0]["evidence"]["page_info_dump"] = str(page_info)
    screen["checklist_evaluations"][0]["evidence"]["wcag_results_dump"] = str(wcag_dump)
    runner.screen_results = [screen]

    report = runner._build_report("testrun01")
    root = Path(generate_checklist_reports(report, tmp_path))
    folder = root / "01_non_text_content"

    assert folder.exists()
    assert (folder / "checklist_01_non_text_content.json").exists()
    assert (folder / "checklist_01_non_text_content.md").exists()
    assert (folder / "issues" / "issue_index.json").exists()
    assert (folder / "states" / "raw" / "state_01_01_initial_load.png").exists()
    assert (folder / "states" / "annotated" / "state_01_01_initial_load_annotated.png").exists()
