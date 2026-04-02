import json
from pathlib import Path

from openpyxl import load_workbook

from accessibility_scanner.agent.cli import APP_CONFIG_MAP
from accessibility_scanner.agent.flow_runner import AgenticFlowRunner
from accessibility_scanner.xlsx_report import REPORT_STANDARD, generate_xlsx_report


class _ManualPage:
    def title(self) -> str:
        return "Client Dashboard"

    def query_selector(self, _selector: str):
        return None


def _sample_screen() -> dict:
    return {
        "label": "otp_screen",
        "url": "https://wsuat.hdfcsec.com/wealthspectrum/portal/client-dashboard/G/100001",
        "unique_key": "screen-key",
        "action_source": "scripted",
        "screenshot": "artifacts/run/otp_screen.png",
        "annotated_screenshot": None,
        "annotation_metadata": None,
        "scroll_screenshots": [],
        "scroll_annotated_screenshots": [],
        "dom_dump": "artifacts/run/otp_screen-dom.html",
        "page_info_dump": "artifacts/run/otp_screen-page-info.json",
        "wcag_results_dump": "artifacts/run/otp_screen-wcag-results.json",
        "wcag_summary_dump": "artifacts/run/otp_screen-wcag-summary.json",
        "wcag_summary": {
            "total_checks": 56,
            "pass": 30,
            "fail": 2,
            "cannot_verify": 20,
            "not_applicable": 4,
            "failures": [],
        },
        "wcag_results": [],
        "contrast_evidence": {},
    }


def test_hdfc_wealthspectrum_config_is_registered() -> None:
    assert APP_CONFIG_MAP["HDFCWSUAT"] == "config/hdfc_wealthspectrum_uat.json"

    config_path = Path(APP_CONFIG_MAP["HDFCWSUAT"])
    config = json.loads(config_path.read_text(encoding="utf-8"))
    assert config["start_url"] == "https://wsuat.hdfcsec.com/wealthspectrum/portal/sign-in"
    assert config["flow_steps"][1]["actions"][0]["type"] == "manual"


def test_hdfc_wealthspectrum_validation_profile_uses_login_id() -> None:
    runner = AgenticFlowRunner(config_path="config/hdfc_wealthspectrum_uat.json", headless=True)
    assert runner.validation_profile["login_id"] == "DUMMYUAT"


def test_pre_login_mode_skips_dashboard_explore() -> None:
    runner = AgenticFlowRunner(
        config_path="config/hdfc_wealthspectrum_uat.json",
        headless=True,
        scan_mode="pre_login",
    )
    login_step, otp_step, explore_step = runner.config["flow_steps"]

    assert runner._should_run_step(login_step) is True
    assert runner._should_run_step(otp_step) is True
    assert runner._should_run_step(explore_step) is False


def test_manual_completion_matches_dashboard_url() -> None:
    runner = AgenticFlowRunner(config_path="config/hdfc_wealthspectrum_uat.json", headless=True)
    runner._page = _ManualPage()

    action = runner.config["flow_steps"][1]["actions"][0]
    dashboard_url = "https://wsuat.hdfcsec.com/wealthspectrum/portal/client-dashboard/G/100001"
    otp_url = "https://wsuat.hdfcsec.com/wealthspectrum/portal/login-otp"

    assert runner._manual_completion_met(action, dashboard_url) is True
    assert runner._manual_completion_met(action, otp_url) is False


def test_build_report_includes_route_log_urls() -> None:
    runner = AgenticFlowRunner(config_path="config/hdfc_wealthspectrum_uat.json", headless=True)
    runner.screen_results = [_sample_screen()]
    runner.visited_urls = ["https://wsuat.hdfcsec.com/wealthspectrum/portal/login-otp"]
    runner.observed_urls = ["https://wsuat.hdfcsec.com/wealthspectrum/portal/sign-in"]
    runner.observed_url_set = set(runner.observed_urls)
    runner.route_log = [
        {
            "event_type": "manual_url_change",
            "source": "manual",
            "from_url": "https://wsuat.hdfcsec.com/wealthspectrum/portal/login-otp",
            "to_url": "https://wsuat.hdfcsec.com/wealthspectrum/portal/client-dashboard/G/100001",
            "target_url": "https://wsuat.hdfcsec.com/wealthspectrum/portal/client-dashboard/G/100001",
            "description": "User enters OTP manually and clicks Sign In",
            "url_changed": True,
        }
    ]

    report = runner._build_report("testrun01")

    assert report["route_log"][0]["event_type"] == "manual_url_change"
    assert "https://wsuat.hdfcsec.com/wealthspectrum/portal/sign-in" in report["urls_visited"]
    assert "https://wsuat.hdfcsec.com/wealthspectrum/portal/login-otp" in report["urls_visited"]
    assert "https://wsuat.hdfcsec.com/wealthspectrum/portal/client-dashboard/G/100001" in report["urls_visited"]


def test_xlsx_route_log_sheet_is_created(tmp_path: Path) -> None:
    report = {
        "run_id": "testrun01",
        "config": "HDFC WealthSpectrum UAT Login + Dashboard",
        "standard": REPORT_STANDARD,
        "screens_analyzed": 1,
        "urls_visited": [
            "https://wsuat.hdfcsec.com/wealthspectrum/portal/sign-in",
            "https://wsuat.hdfcsec.com/wealthspectrum/portal/login-otp",
            "https://wsuat.hdfcsec.com/wealthspectrum/portal/client-dashboard/G/100001",
        ],
        "totals": {"pass": 30, "fail": 2, "cannot_verify": 20},
        "all_failures": [],
        "cannot_verify_policy": "pass_leaning",
        "cannot_verify_threshold": 31,
        "cannot_verify_enforcement": "both",
        "cannot_verify_metrics": {
            "checkpoint_count": 0,
            "instance_count": 0,
            "threshold": 31,
            "enforcement": "both",
            "checkpoint_within_threshold": True,
            "instance_within_threshold": True,
            "within_threshold": True,
        },
        "route_log": [
            {
                "event_type": "scripted_click",
                "source": "scripted",
                "from_url": "https://wsuat.hdfcsec.com/wealthspectrum/portal/sign-in",
                "to_url": "https://wsuat.hdfcsec.com/wealthspectrum/portal/login-otp",
                "target_url": "",
                "description": "Click Sign In on the login page",
                "element_text": "",
                "url_changed": True,
            }
        ],
        "screens": [_sample_screen()],
    }

    xlsx_path = tmp_path / "wealthspectrum.xlsx"
    generate_xlsx_report(report, str(xlsx_path))
    wb = load_workbook(str(xlsx_path))

    assert "Route Log" in wb.sheetnames
    ws_routes = wb["Route Log"]
    assert ws_routes.cell(row=2, column=2).value == "scripted_click"
    assert ws_routes.cell(row=2, column=4).value == "https://wsuat.hdfcsec.com/wealthspectrum/portal/sign-in"
