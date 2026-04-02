"""Workbook-backed checklist metadata for detailed WCAG reporting."""
from __future__ import annotations

from dataclasses import asdict, dataclass
from functools import lru_cache
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


_DEFAULT_WORKBOOK_PATH = Path(__file__).resolve().parents[2] / "wcag_web_agent_checklist.xlsx"
_HEADER_ALIASES = {
    "principle": "principle",
    "guideline": "guideline",
    "sc id": "sc_id",
    "sc title": "sc_title",
    "level": "level",
    "requirement summary": "requirement_summary",
    "automated agent goal": "automated_agent_goal",
    "preconditions / test data": "preconditions_test_data",
    "automated agent actions": "automated_agent_actions",
    "dom targets / signals": "dom_targets_signals",
    "required evidence for llm": "required_evidence_for_llm",
    "machine pass criteria": "machine_pass_criteria",
    "failure heuristics / flags": "failure_heuristics_flags",
    "exceptions / notes": "exceptions_notes",
    "suggested tooling": "suggested_tooling",
    "output json fields": "output_json_fields",
    "notes": "notes",
}
_REQUIRED_FIELDS = {
    "principle",
    "guideline",
    "sc_id",
    "sc_title",
    "level",
    "requirement_summary",
    "automated_agent_goal",
    "preconditions_test_data",
    "automated_agent_actions",
    "dom_targets_signals",
    "required_evidence_for_llm",
    "machine_pass_criteria",
    "failure_heuristics_flags",
    "exceptions_notes",
    "suggested_tooling",
    "output_json_fields",
}


@dataclass(frozen=True)
class ChecklistSpec:
    sequence: int
    principle: str
    guideline: str
    sc_id: str
    sc_title: str
    level: str
    requirement_summary: str
    automated_agent_goal: str
    preconditions_test_data: str
    automated_agent_actions: str
    dom_targets_signals: str
    required_evidence_for_llm: str
    machine_pass_criteria: str
    failure_heuristics_flags: str
    exceptions_notes: str
    suggested_tooling: str
    output_json_fields: str
    notes: str = ""

    @property
    def slug(self) -> str:
        return _slugify(self.sc_title)

    @property
    def output_json_field_list(self) -> list[str]:
        return [
            item.strip()
            for item in re.split(r",|\n", self.output_json_fields or "")
            if item and item.strip()
        ]

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["slug"] = self.slug
        data["output_json_field_list"] = self.output_json_field_list
        return data


def _slugify(text: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", (text or "").strip().lower())
    return slug.strip("_") or "checklist"


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_headers(raw_headers: list[Any]) -> dict[int, str]:
    normalized: dict[int, str] = {}
    for idx, value in enumerate(raw_headers, start=1):
        key = _clean_text(value).lower()
        if not key:
            continue
        mapped = _HEADER_ALIASES.get(key)
        if mapped:
            normalized[idx] = mapped
    missing = sorted(_REQUIRED_FIELDS - set(normalized.values()))
    if missing:
        raise ValueError(f"Checklist workbook is missing required columns: {missing}")
    return normalized


@lru_cache(maxsize=1)
def load_checklist_specs(workbook_path: str | None = None) -> tuple[ChecklistSpec, ...]:
    workbook = Path(workbook_path) if workbook_path else _DEFAULT_WORKBOOK_PATH
    wb = load_workbook(workbook, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_map = _normalize_headers([ws.cell(1, c).value for c in range(1, ws.max_column + 1)])

    specs: list[ChecklistSpec] = []
    for row_idx in range(2, ws.max_row + 1):
        values = {
            field_name: _clean_text(ws.cell(row_idx, col_idx).value)
            for col_idx, field_name in header_map.items()
        }
        sc_id = values.get("sc_id", "")
        if not sc_id:
            continue
        specs.append(
            ChecklistSpec(
                sequence=len(specs) + 1,
                principle=values.get("principle", ""),
                guideline=values.get("guideline", ""),
                sc_id=sc_id,
                sc_title=values.get("sc_title", ""),
                level=values.get("level", ""),
                requirement_summary=values.get("requirement_summary", ""),
                automated_agent_goal=values.get("automated_agent_goal", ""),
                preconditions_test_data=values.get("preconditions_test_data", ""),
                automated_agent_actions=values.get("automated_agent_actions", ""),
                dom_targets_signals=values.get("dom_targets_signals", ""),
                required_evidence_for_llm=values.get("required_evidence_for_llm", ""),
                machine_pass_criteria=values.get("machine_pass_criteria", ""),
                failure_heuristics_flags=values.get("failure_heuristics_flags", ""),
                exceptions_notes=values.get("exceptions_notes", ""),
                suggested_tooling=values.get("suggested_tooling", ""),
                output_json_fields=values.get("output_json_fields", ""),
                notes=values.get("notes", ""),
            )
        )
    return tuple(specs)


@lru_cache(maxsize=1)
def load_checklist_spec_map(workbook_path: str | None = None) -> dict[str, ChecklistSpec]:
    return {spec.sc_id: spec for spec in load_checklist_specs(workbook_path)}

