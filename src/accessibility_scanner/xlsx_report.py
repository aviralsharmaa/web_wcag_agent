"""Generate a structured XLSX report for WCAG A/AA (Expanded) audit results."""
from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from .checkpoints import CHECKPOINTS


# ── Expanded WCAG A/AA guideline metadata (56 criteria) ──────────────────
REPORT_STANDARD = "WCAG A/AA (Expanded)"
WCAG_GUIDELINES: list[dict[str, str]] = [
    {"id": "1.1.1", "title": "Non-text Content", "level": "A",
     "requirement": "All non-text content has a text alternative that serves the equivalent purpose."},
    {"id": "1.2.1", "title": "Audio-only and Video-only (Prerecorded)", "level": "A",
     "requirement": "An alternative is provided for prerecorded audio-only and prerecorded video-only media."},
    {"id": "1.2.2", "title": "Captions (Prerecorded)", "level": "A",
     "requirement": "Captions are provided for all prerecorded audio content in synchronized media."},
    {"id": "1.2.3", "title": "Audio Description or Media Alternative", "level": "A",
     "requirement": "An alternative for time-based media or audio description of prerecorded video content is provided."},
    {"id": "1.2.4", "title": "Captions (Live)", "level": "AA",
     "requirement": "Captions are provided for all live audio content in synchronized media."},
    {"id": "1.2.5", "title": "Audio Description (Prerecorded)", "level": "AA",
     "requirement": "Audio description is provided for all prerecorded video content in synchronized media."},
    {"id": "1.3.1", "title": "Info and Relationships", "level": "A",
     "requirement": "Information, structure, and relationships conveyed through presentation can be programmatically determined."},
    {"id": "1.3.2", "title": "Meaningful Sequence", "level": "A",
     "requirement": "When the sequence in which content is presented affects its meaning, a correct reading sequence can be programmatically determined."},
    {"id": "1.3.3", "title": "Sensory Characteristics", "level": "A",
     "requirement": "Instructions do not rely solely on sensory characteristics such as shape, color, size, visual location, orientation, or sound."},
    {"id": "1.3.4", "title": "Orientation", "level": "AA",
     "requirement": "Content does not restrict its view and operation to a single display orientation."},
    {"id": "1.3.5", "title": "Identify Input Purpose", "level": "AA",
     "requirement": "The purpose of input fields collecting user information can be programmatically determined."},
    {"id": "1.4.1", "title": "Use of Color", "level": "A",
     "requirement": "Color is not used as the only visual means of conveying information."},
    {"id": "1.4.2", "title": "Audio Control", "level": "A",
     "requirement": "A mechanism is available to pause or stop audio that plays automatically for more than 3 seconds."},
    {"id": "1.4.3", "title": "Contrast (Minimum)", "level": "AA",
     "requirement": "Text and images of text have a contrast ratio of at least 4.5:1 (3:1 for large text)."},
    {"id": "1.4.4", "title": "Resize Text", "level": "AA",
     "requirement": "Text can be resized without assistive technology up to 200% without loss of content or functionality."},
    {"id": "1.4.5", "title": "Images of Text", "level": "AA",
     "requirement": "Text is used to convey information rather than images of text wherever possible."},
    {"id": "1.4.10", "title": "Reflow", "level": "AA",
     "requirement": "Content can be presented without loss of information or functionality at 320 CSS px width / 256 CSS px height."},
    {"id": "1.4.11", "title": "Non-text Contrast", "level": "AA",
     "requirement": "UI components and graphical objects have a contrast ratio of at least 3:1."},
    {"id": "1.4.12", "title": "Text Spacing", "level": "AA",
     "requirement": "No loss of content or functionality occurs when text spacing is adjusted."},
    {"id": "1.4.13", "title": "Content on Hover or Focus", "level": "AA",
     "requirement": "Content triggered by pointer hover or keyboard focus is dismissible, hoverable, and persistent."},
    {"id": "2.1.1", "title": "Keyboard", "level": "A",
     "requirement": "All functionality is operable through a keyboard interface without requiring specific timings."},
    {"id": "2.1.2", "title": "No Keyboard Trap", "level": "A",
     "requirement": "The user can move focus away from any component using only a keyboard interface."},
    {"id": "2.1.4", "title": "Character Key Shortcuts", "level": "A",
     "requirement": "If character key shortcuts exist, they can be turned off, remapped, or are only active when the component has focus."},
    {"id": "2.2.1", "title": "Timing Adjustable", "level": "A",
     "requirement": "For time limits, the user can turn off, adjust, or extend the time limit."},
    {"id": "2.2.2", "title": "Pause, Stop, Hide", "level": "A",
     "requirement": "Moving, blinking, scrolling, or auto-updating content can be paused, stopped, or hidden."},
    {"id": "2.3.1", "title": "Three Flashes or Below Threshold", "level": "A",
     "requirement": "Pages do not contain anything that flashes more than three times per second."},
    {"id": "2.4.1", "title": "Bypass Blocks", "level": "A",
     "requirement": "A mechanism is available to bypass blocks of content that are repeated on multiple pages."},
    {"id": "2.4.2", "title": "Page Titled", "level": "A",
     "requirement": "Web pages have titles that describe topic or purpose."},
    {"id": "2.4.3", "title": "Focus Order", "level": "A",
     "requirement": "Focusable components receive focus in an order that preserves meaning and operability."},
    {"id": "2.4.4", "title": "Link Purpose (In Context)", "level": "A",
     "requirement": "The purpose of each link can be determined from the link text alone or together with its context."},
    {"id": "2.4.5", "title": "Multiple Ways", "level": "AA",
     "requirement": "More than one way is available to locate a web page within a set of web pages."},
    {"id": "2.4.6", "title": "Headings and Labels", "level": "AA",
     "requirement": "Headings and labels describe topic or purpose."},
    {"id": "2.4.7", "title": "Focus Visible", "level": "AA",
     "requirement": "Any keyboard-operable user interface has a mode of operation where the keyboard focus indicator is visible."},
    {"id": "2.4.11", "title": "Focus Not Obscured (Minimum)", "level": "AA",
     "requirement": "When a UI component receives keyboard focus, it is not entirely hidden by author-created content."},
    {"id": "2.5.1", "title": "Pointer Gestures", "level": "A",
     "requirement": "All functionality that uses multipoint or path-based gestures can be operated with a single pointer."},
    {"id": "2.5.2", "title": "Pointer Cancellation", "level": "A",
     "requirement": "For single-pointer functionality, at least one of: the down-event is not used, completion is on up-event, or actions can be aborted/undone."},
    {"id": "2.5.3", "title": "Label in Name", "level": "A",
     "requirement": "For UI components with labels that include text or images of text, the accessible name contains the text that is presented visually."},
    {"id": "2.5.4", "title": "Motion Actuation", "level": "A",
     "requirement": "Functionality triggered by device motion can be operated via a user interface component, and motion triggering can be disabled."},
    {"id": "2.5.7", "title": "Dragging Movements", "level": "AA",
     "requirement": "Functionality achievable with a dragging movement can be achieved with a single pointer without dragging."},
    {"id": "2.5.8", "title": "Target Size (Minimum)", "level": "AA",
     "requirement": "The size of the target for pointer inputs is at least 24 by 24 CSS pixels."},
    {"id": "3.1.1", "title": "Language of Page", "level": "A",
     "requirement": "The default human language of each web page can be programmatically determined."},
    {"id": "3.1.2", "title": "Language of Parts", "level": "AA",
     "requirement": "The human language of each passage or phrase in the content can be programmatically determined."},
    {"id": "3.2.1", "title": "On Focus", "level": "A",
     "requirement": "When any UI component receives focus, it does not initiate a change of context."},
    {"id": "3.2.2", "title": "On Input", "level": "A",
     "requirement": "Changing the setting of any UI component does not automatically cause a change of context unless the user has been advised."},
    {"id": "3.2.3", "title": "Consistent Navigation", "level": "AA",
     "requirement": "Navigational mechanisms that are repeated on multiple pages occur in the same relative order."},
    {"id": "3.2.4", "title": "Consistent Identification", "level": "AA",
     "requirement": "Components that have the same functionality are identified consistently."},
    {"id": "3.2.6", "title": "Consistent Help", "level": "A",
     "requirement": "Help mechanisms occur in the same relative order on multiple pages."},
    {"id": "3.3.1", "title": "Error Identification", "level": "A",
     "requirement": "If an input error is automatically detected, the item that is in error is identified and described in text."},
    {"id": "3.3.2", "title": "Labels or Instructions", "level": "A",
     "requirement": "Labels or instructions are provided when content requires user input."},
    {"id": "3.3.3", "title": "Error Suggestion", "level": "AA",
     "requirement": "If an input error is detected and suggestions are known, they are provided to the user."},
    {"id": "3.3.4", "title": "Error Prevention (Legal, Financial, Data)", "level": "AA",
     "requirement": "For pages with legal commitments or financial transactions, submissions are reversible, checked, or confirmed."},
    {"id": "3.3.7", "title": "Redundant Entry", "level": "A",
     "requirement": "Information previously entered by or provided to the user is auto-populated or available for selection."},
    {"id": "3.3.8", "title": "Accessible Authentication (Minimum)", "level": "AA",
     "requirement": "A cognitive function test is not required for any step in an authentication process unless an alternative is provided."},
    {"id": "4.1.1", "title": "Parsing", "level": "A",
     "requirement": "Elements have complete start/end tags, are properly nested, and contain no duplicate attributes/IDs that break parsing."},
    {"id": "4.1.2", "title": "Name, Role, Value", "level": "A",
     "requirement": "For all UI components, the name and role can be programmatically determined; states, properties, and values can be programmatically set."},
    {"id": "4.1.3", "title": "Status Messages", "level": "AA",
     "requirement": "Status messages can be programmatically determined through role or properties so they can be presented by assistive technologies."},
]

_GUIDELINE_MAP = {g["id"]: g for g in WCAG_GUIDELINES}

# ── Styling constants ────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
_FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
_CV_FILL = PatternFill("solid", fgColor="FFEB9C")
_NA_FILL = PatternFill("solid", fgColor="D9E2F3")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_STATUS_FILLS = {
    "Pass": _PASS_FILL,
    "Fail": _FAIL_FILL,
    "Cannot verify automatically": _CV_FILL,
    "Not applicable": _NA_FILL,
}


def _style_header_row(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _auto_width(ws, min_width: int = 12, max_width: int = 55):
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = max_len + 2


def _validate_guideline_sync() -> None:
    runtime_ids = {meta.checkpoint_id for meta in CHECKPOINTS}
    guideline_ids = {item["id"] for item in WCAG_GUIDELINES}
    missing = sorted(runtime_ids - guideline_ids)
    extra = sorted(guideline_ids - runtime_ids)
    if missing or extra:
        raise ValueError(
            f"Guideline/checkpoint mismatch. Missing in XLSX metadata: {missing}. Extra in XLSX metadata: {extra}."
        )


def _reduce_status(statuses: list[str]) -> str:
    if not statuses:
        return "Not evaluated"
    if "Fail" in statuses:
        return "Fail"
    if "Cannot verify automatically" in statuses:
        return "Cannot verify automatically"
    if all(item == "Not applicable" for item in statuses):
        return "Not applicable"
    if "Pass" in statuses:
        return "Pass"
    return "Not evaluated"


def _aggregate_checkpoint_rows(report: dict[str, Any]) -> dict[str, dict[str, Any]]:
    grouped: dict[str, list[dict[str, str]]] = {}
    for screen in report.get("screens", []):
        screen_url = screen.get("url", "")
        for row in screen.get("wcag_results", []):
            checkpoint_id = row.get("checkpoint_id", "")
            if not checkpoint_id:
                continue
            grouped.setdefault(checkpoint_id, []).append(
                {
                    "status": row.get("status", ""),
                    "rationale": row.get("rationale", ""),
                    "page": screen_url,
                }
            )

    aggregate: dict[str, dict[str, Any]] = {}
    for checkpoint_id, rows in grouped.items():
        statuses = [r["status"] for r in rows]
        resolved_status = _reduce_status(statuses)

        rationale = ""
        for row in rows:
            if row["status"] == resolved_status and row["rationale"]:
                rationale = row["rationale"]
                break
        if not rationale:
            for row in rows:
                if row["rationale"]:
                    rationale = row["rationale"]
                    break

        pages = []
        for row in rows:
            if row["status"] == resolved_status and row["page"] not in pages:
                pages.append(row["page"])
        if not pages:
            for row in rows:
                if row["page"] not in pages:
                    pages.append(row["page"])

        aggregate[checkpoint_id] = {
            "status": resolved_status,
            "rationale": rationale,
            "pages": pages,
        }
    return aggregate


def generate_xlsx_report(report: dict[str, Any], output_path: str) -> str:
    """Generate a structured XLSX from the agentic scan report dict.

    Args:
        report: The JSON-serializable report dict produced by AgenticFlowRunner._build_report().
        output_path: Absolute or relative path for the output .xlsx file.

    Returns:
        The output_path string.
    """
    _validate_guideline_sync()
    wb = Workbook()
    checkpoint_agg = _aggregate_checkpoint_rows(report)

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Metric", "Value"])
    _style_header_row(ws_summary, 2)

    totals = report.get("totals", {})
    cv_instance_total = totals.get("cannot_verify", 0)
    cv_checkpoint_total = sum(
        1 for item in checkpoint_agg.values() if item.get("status") == "Cannot verify automatically"
    )
    cv_threshold = (
        report.get("cannot_verify_metrics", {}).get("threshold")
        or report.get("cannot_verify_threshold")
        or 31
    )
    cv_checkpoint_ok = cv_checkpoint_total <= cv_threshold
    cv_instance_ok = cv_instance_total <= cv_threshold
    cv_enforcement = (
        report.get("cannot_verify_metrics", {}).get("enforcement")
        or report.get("cannot_verify_enforcement")
        or "both"
    )
    if cv_enforcement == "checkpoint":
        cv_within = cv_checkpoint_ok
    elif cv_enforcement == "instance":
        cv_within = cv_instance_ok
    else:
        cv_within = cv_checkpoint_ok and cv_instance_ok

    ws_summary.append(["Run ID", report.get("run_id", "")])
    ws_summary.append(["Config", report.get("config", "")])
    ws_summary.append(["Standard", report.get("standard", REPORT_STANDARD)])
    ws_summary.append(["Scan Mode", report.get("scan_mode", "full_scan")])
    ws_summary.append(["Cannot Verify Policy", report.get("cannot_verify_policy", "pass_leaning")])
    ws_summary.append(["Cannot Verify Enforcement", cv_enforcement])
    ws_summary.append(["Screens Analyzed", report.get("screens_analyzed", 0)])
    ws_summary.append(["URLs Visited", len(report.get("urls_visited", []))])
    ws_summary.append(["Total PASS", totals.get("pass", 0)])
    ws_summary.append(["Total FAIL", totals.get("fail", 0)])
    ws_summary.append(["Total CANNOT VERIFY (Instances)", cv_instance_total])
    ws_summary.append(["Total CANNOT VERIFY (Checkpoints)", cv_checkpoint_total])
    ws_summary.append(["CV Threshold (<=)", cv_threshold])
    ws_summary.append(["CV Threshold Checkpoints", "PASS" if cv_checkpoint_ok else "FAIL"])
    ws_summary.append(["CV Threshold Instances", "PASS" if cv_instance_ok else "FAIL"])
    ws_summary.append(["CV Threshold Overall", "PASS" if cv_within else "FAIL"])
    ws_summary.append(["Unique Failures", len(report.get("all_failures", []))])

    for row in range(2, ws_summary.max_row + 1):
        metric = ws_summary.cell(row=row, column=1).value
        if metric in {"CV Threshold Checkpoints", "CV Threshold Instances", "CV Threshold Overall"}:
            value_cell = ws_summary.cell(row=row, column=2)
            value = str(value_cell.value or "")
            if value == "PASS":
                value_cell.fill = _PASS_FILL
            elif value == "FAIL":
                value_cell.fill = _FAIL_FILL
            value_cell.alignment = Alignment(horizontal="center")
    _auto_width(ws_summary)

    # ── Sheet 2: WCAG Guidelines ──────────────────────────────────────
    ws_wcag = wb.create_sheet("WCAG Guidelines")
    headers = [
        "SC ID", "Title", "Level", "Requirement",
        "Status", "Rationale", "Page URL", "Evidence",
    ]
    ws_wcag.append(headers)
    _style_header_row(ws_wcag, len(headers))

    # Build a lookup from full per-screen wcag_results.

    for guideline in WCAG_GUIDELINES:
        gid = guideline["id"]
        agg = checkpoint_agg.get(gid)
        if agg:
            status = agg["status"]
            rationale = agg["rationale"]
            pages = "; ".join(agg["pages"][:5])
        else:
            status = "Not evaluated"
            rationale = ""
            pages = ""

        row_num = ws_wcag.max_row + 1
        ws_wcag.append([
            gid,
            guideline["title"],
            guideline["level"],
            guideline["requirement"],
            status,
            rationale,
            pages,
            "",  # Evidence column (screenshots etc.)
        ])

        # Color the status cell
        status_cell = ws_wcag.cell(row=row_num, column=5)
        fill = _STATUS_FILLS.get(status)
        if fill:
            status_cell.fill = fill

    # Wrap text on requirement and rationale columns
    for row in ws_wcag.iter_rows(min_row=2, max_row=ws_wcag.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _THIN_BORDER

    _auto_width(ws_wcag)

    # ── Sheet 3: Per-Screen Results ───────────────────────────────────
    ws_screens = wb.create_sheet("Per-Screen Results")
    screen_headers = [
        "#", "Screen Label", "URL", "Screenshot",
        "Total Pass", "Total Fail", "Cannot Verify",
        "Top Failures",
    ]
    ws_screens.append(screen_headers)
    _style_header_row(ws_screens, len(screen_headers))

    for i, screen in enumerate(report.get("screens", []), 1):
        s = screen.get("wcag_summary", {})
        failures_str = "; ".join(
            f"[{f['checkpoint']}] {f.get('rationale', '')[:60]}"
            for f in s.get("failures", [])[:5]
        )
        ws_screens.append([
            i,
            screen.get("label", ""),
            screen.get("url", ""),
            screen.get("screenshot", ""),
            s.get("pass", 0),
            s.get("fail", 0),
            s.get("cannot_verify", 0),
            failures_str,
        ])

    for row in ws_screens.iter_rows(min_row=2, max_row=ws_screens.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _THIN_BORDER

    _auto_width(ws_screens)

    checklist_rollup = report.get("checklist_rollup", [])
    if checklist_rollup:
        ws_checklists = wb.create_sheet("Checklist Detail")
        checklist_headers = [
            "SC ID",
            "Title",
            "Level",
            "Aggregate Status",
            "Screens",
            "Pages",
            "Agent Goal",
            "Required Evidence For LLM",
            "Machine Pass Criteria",
            "Failure Heuristics / Flags",
        ]
        ws_checklists.append(checklist_headers)
        _style_header_row(ws_checklists, len(checklist_headers))

        for item in checklist_rollup:
            row_num = ws_checklists.max_row + 1
            ws_checklists.append(
                [
                    item.get("sc_id", ""),
                    item.get("sc_title", ""),
                    item.get("level", ""),
                    item.get("aggregate_status", ""),
                    len(item.get("screen_evaluations", [])),
                    "; ".join(item.get("pages", [])[:5]),
                    item.get("automated_agent_goal", ""),
                    item.get("required_evidence_for_llm", ""),
                    item.get("machine_pass_criteria", ""),
                    item.get("failure_heuristics_flags", ""),
                ]
            )
            fill = _STATUS_FILLS.get(item.get("aggregate_status", ""))
            if fill:
                ws_checklists.cell(row=row_num, column=4).fill = fill

        for row in ws_checklists.iter_rows(min_row=2, max_row=ws_checklists.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = _THIN_BORDER
        _auto_width(ws_checklists)

    # ── Sheet 4: Route Log ────────────────────────────────────────────
    route_log = report.get("route_log", [])
    if route_log:
        ws_routes = wb.create_sheet("Route Log")
        route_headers = [
            "#", "Event Type", "Source", "From URL", "To URL",
            "Target URL", "Description", "Element Text", "Changed URL",
        ]
        ws_routes.append(route_headers)
        _style_header_row(ws_routes, len(route_headers))

        for idx, event in enumerate(route_log, 1):
            ws_routes.append([
                idx,
                event.get("event_type", ""),
                event.get("source", ""),
                event.get("from_url", ""),
                event.get("to_url", ""),
                event.get("target_url", ""),
                event.get("description", ""),
                event.get("element_text", ""),
                "Yes" if event.get("url_changed") else "No",
            ])

        for row in ws_routes.iter_rows(min_row=2, max_row=ws_routes.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = _THIN_BORDER

        _auto_width(ws_routes)

    # ── Save ──────────────────────────────────────────────────────────
    wb.save(output_path)
    return output_path
